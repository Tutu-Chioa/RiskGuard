from flask import Flask, jsonify, request, send_from_directory, send_file
from datetime import datetime, timedelta
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
from functools import wraps
import os
import sys
import json
import subprocess
import secrets
import logging
import time

# 生产环境不向前端暴露详细错误，仅记录日志
_is_production = os.environ.get('FLASK_ENV') == 'production' or os.environ.get('SKIP_DROP_TABLES', '').strip() in ('1', 'true', 'yes')
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# 从 backend/.env 加载环境变量（MEDIACRAWLER_PATH、SECRET_KEY、SMTP 等），无需每次手动 export
_backend_dir = os.path.dirname(os.path.abspath(__file__))
_project_root = os.path.dirname(_backend_dir)
# 开发/非打包时：确保项目根在 path 最前，否则 from backend.services.xxx 会报 No module named 'backend'
if not getattr(sys, 'frozen', False) and _project_root not in sys.path:
    sys.path.insert(0, _project_root)
# .app 下再次确保含 backend 的目录在 path 最前（run_desktop 已设，此处兜底，避免 no module named 'services'）
if getattr(sys, 'frozen', False):
    _exe_dir = os.path.dirname(os.path.abspath(sys.executable))
    _frameworks = os.path.abspath(os.path.join(_exe_dir, '..', 'Frameworks'))
    if _frameworks not in sys.path:
        sys.path.insert(0, _frameworks)
    _meipass = getattr(sys, '_MEIPASS', '')
    if _meipass and _meipass not in sys.path:
        sys.path.insert(0, _meipass)
# 打包为 Mac .app 或 PyInstaller 时使用用户目录存放数据库与上传，否则使用项目 data/
_standalone_data = getattr(sys, 'frozen', False) or os.environ.get('STANDALONE_APP', '').strip() in ('1', 'true', 'yes')
if _standalone_data:
    _DATA_DIR = os.path.join(os.path.expanduser('~'), 'Library', 'Application Support', 'RiskGuard')
    os.makedirs(_DATA_DIR, exist_ok=True)
    os.environ['RISKGUARD_DATA_DIR'] = _DATA_DIR
else:
    _DATA_DIR = os.path.join(_project_root, 'data')
_env_path = os.path.join(_backend_dir, '.env')
# 数据库路径与连接：统一 timeout 减少 "database is locked"
_DB_PATH = os.path.join(_DATA_DIR, 'risk_platform.db')
def _db_conn(path=None, timeout=30):
    return sqlite3.connect(path or _DB_PATH, timeout=timeout)
try:
    from dotenv import load_dotenv
    load_dotenv(_env_path)
except ImportError:
    pass
# 若未安装 python-dotenv 或 .env 中 MEDIACRAWLER_PATH 仍为空，尝试手动读 backend/.env 中 MEDIACRAWLER_PATH
if not os.environ.get('MEDIACRAWLER_PATH') and os.path.isfile(_env_path):
    try:
        with open(_env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('MEDIACRAWLER_PATH=') and not line.startswith('#'):
                    val = line.split('=', 1)[1].strip().strip('"\'')
                    if val:
                        os.environ['MEDIACRAWLER_PATH'] = val
                    break
    except Exception:
        pass
# 若仍为空，自动使用「与 backend 同级的 MediaCrawler」目录（常见：把 MediaCrawler 放在 sys2 里）
if not os.environ.get('MEDIACRAWLER_PATH'):
    _project_root = os.path.dirname(_backend_dir)
    for _name in ('MediaCrawler', 'MediaCrawler-main'):
        _auto_mc = os.path.join(_project_root, _name)
        if os.path.isfile(os.path.join(_auto_mc, 'main.py')):
            os.environ['MEDIACRAWLER_PATH'] = _auto_mc
            break
# .app/standalone 下从应用数据目录加载 .env（用户可在设置页保存 MEDIACRAWLER_PATH，写入此处）
if _standalone_data:
    _data_env = os.path.join(_DATA_DIR, '.env')
    if os.path.isfile(_data_env):
        try:
            from dotenv import load_dotenv
            load_dotenv(_data_env)
        except Exception:
            pass
    if not os.environ.get('MEDIACRAWLER_PATH') and os.path.isfile(_data_env):
        try:
            with open(_data_env, 'r', encoding='utf-8') as _f:
                for _line in _f:
                    _line = _line.strip()
                    if _line.startswith('MEDIACRAWLER_PATH=') and not _line.startswith('#'):
                        _val = _line.split('=', 1)[1].strip().strip('"\'')
                        if _val:
                            os.environ['MEDIACRAWLER_PATH'] = _val
                        break
        except Exception:
            pass
# .app 内嵌 MediaCrawler：frozen 时强制在首次运行把 bundle 中的 MediaCrawler 复制到应用数据目录，实现「安装即用」
if getattr(sys, 'frozen', False):
    _copied_mc = os.path.join(_DATA_DIR, 'MediaCrawler')
    _need_copy = not os.path.isfile(os.path.join(_copied_mc, 'main.py'))
    _bundled_mc = None
    _exe_dir = os.path.dirname(os.path.abspath(sys.executable))

    def _check_mc(p):
        if not p or not os.path.isdir(p):
            return False
        # PyInstaller 有时把 main.py 打成目录，其内才有 main.py 文件
        main_py = os.path.join(p, 'main.py')
        return os.path.isfile(main_py) or os.path.isfile(os.path.join(main_py, 'main.py'))
    # 1) 用 __file__ 定位：app.py 在 .../Frameworks/backend/app.py，同级目录为 Frameworks，其下 MediaCrawler
    _app_file_dir = os.path.dirname(os.path.abspath(__file__))
    _frameworks_from_file = os.path.join(os.path.dirname(_app_file_dir), 'MediaCrawler')
    if _check_mc(_frameworks_from_file):
        _bundled_mc = os.path.abspath(_frameworks_from_file)
    # 2) 可执行文件在 Contents/MacOS，MediaCrawler 在 Contents/Frameworks
    if not _bundled_mc:
        _frameworks_mc = os.path.abspath(os.path.join(_exe_dir, '..', 'Frameworks', 'MediaCrawler'))
        if _check_mc(_frameworks_mc):
            _bundled_mc = _frameworks_mc
    # 3) PyInstaller onedir：_MEIPASS 或可执行文件同目录
    if not _bundled_mc:
        _meipass = getattr(sys, '_MEIPASS', '')
        if _meipass:
            _m = os.path.join(_meipass, 'MediaCrawler')
            if _check_mc(_m):
                _bundled_mc = os.path.abspath(_m)
    if not _bundled_mc:
        _exe_sibling = os.path.join(_exe_dir, 'MediaCrawler')
        if _check_mc(_exe_sibling):
            _bundled_mc = os.path.abspath(_exe_sibling)
    def _flatten_mc_dirs(root):
        """PyInstaller 把路径打成「目录+同名文件」(如 database/db.py/ 内只有 db.py)，展平为正常文件。"""
        try:
            import shutil
            to_replace = []
            for dirpath, dirs, files in os.walk(root, topdown=False):
                for d in dirs:
                    p = os.path.join(dirpath, d)
                    inner = os.path.join(p, d)
                    if os.path.isfile(inner) and len(os.listdir(p)) == 1:
                        to_replace.append((p, inner))
            for parent_dir, inner_file in to_replace:
                try:
                    with open(inner_file, 'rb') as f:
                        data = f.read()
                    shutil.rmtree(parent_dir)
                    with open(parent_dir, 'wb') as f:
                        f.write(data)
                except Exception:
                    pass
        except Exception:
            pass
    if _bundled_mc and _need_copy:
        try:
            import shutil
            os.makedirs(_DATA_DIR, exist_ok=True)
            if os.path.exists(_copied_mc):
                shutil.rmtree(_copied_mc)
            shutil.copytree(_bundled_mc, _copied_mc, symlinks=False, ignore=shutil.ignore_patterns('__pycache__', '*.pyc', '.DS_Store'))
            _flatten_mc_dirs(_copied_mc)
        except Exception:
            pass
    def _has_main(p):
        if not p:
            return False
        mp = os.path.join(p, 'main.py')
        return os.path.isfile(mp) or os.path.isfile(os.path.join(mp, 'main.py'))
    # 已存在的复制目录也可能是未展平结构，每次启动时展平一次（幂等）
    if _bundled_mc and os.path.isdir(_copied_mc):
        _flatten_mc_dirs(_copied_mc)
    # 优先用已复制到应用数据目录的路径（可写、持久）；若复制未成功则用 bundle 内路径，保证「已配置」可用
    if _bundled_mc and os.path.isdir(_copied_mc) and _has_main(_copied_mc):
        os.environ['MEDIACRAWLER_PATH'] = _copied_mc
    elif _bundled_mc:
        os.environ['MEDIACRAWLER_PATH'] = _bundled_mc
    # 调试：写入解析结果，便于排查「未配置」（需用 .app 启动才会生成）
    try:
        os.makedirs(_DATA_DIR, exist_ok=True)
        _log_path = os.path.join(_DATA_DIR, 'mediacrawler_resolve.log')
        with open(_log_path, 'a', encoding='utf-8') as _lf:
            _lf.write('%s frozen=1 exe_dir=%r __file__=%r bundled_mc=%r copied_ok=%s MEDIACRAWLER_PATH=%r\n' % (
                datetime.now().isoformat(), _exe_dir, __file__, _bundled_mc,
                _has_main(_copied_mc) if _copied_mc else False,
                os.environ.get('MEDIACRAWLER_PATH', '')
            ))
    except Exception:
        pass
import smtplib
import ssl
import threading
import base64
import pyotp
import qrcode
import io
import time
import asyncio
from concurrent.futures import ThreadPoolExecutor

# 导入CORS
from flask_cors import CORS

app = Flask(__name__)
# 优先从环境变量读取密钥，避免生产环境硬编码
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')

# 邮件发送：优先从管理后台配置（system_settings 表）读取 SMTP，若无则从环境变量读取
def _get_smtp_config():
    """返回 SMTP 配置 dict：host, port, user, password, from_addr。优先数据库，其次环境变量。"""
    db_path = _DB_PATH
    out = {'host': '', 'port': 587, 'user': '', 'password': '', 'from_addr': ''}
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        for key in ('smtp_host', 'smtp_port', 'smtp_user', 'smtp_password', 'smtp_from'):
            cursor.execute("SELECT value FROM system_settings WHERE key=?", (key,))
            row = cursor.fetchone()
            if row and row[0] is not None and str(row[0]).strip():
                if key == 'smtp_port':
                    try:
                        out['port'] = int(row[0])
                    except ValueError:
                        out['port'] = 587
                elif key == 'smtp_host':
                    out['host'] = str(row[0]).strip()
                elif key == 'smtp_user':
                    out['user'] = str(row[0]).strip()
                elif key == 'smtp_password':
                    out['password'] = str(row[0]).strip()
                elif key == 'smtp_from':
                    out['from_addr'] = str(row[0]).strip()
        conn.close()
    except Exception:
        pass
    if not out['host'] or not out['user']:
        out['host'] = os.environ.get('SMTP_HOST', '')
        out['port'] = int(os.environ.get('SMTP_PORT', '587'))
        out['user'] = os.environ.get('SMTP_USER', '')
        out['password'] = os.environ.get('SMTP_PASSWORD', '')
        out['from_addr'] = (os.environ.get('SMTP_FROM') or out['user'] or '').strip()
    if not out['from_addr'] and out['user']:
        out['from_addr'] = out['user']
    return out

def _smtp_configured():
    c = _get_smtp_config()
    return bool(c.get('host') and c.get('user'))

def _get_user_email(user_id):
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT email FROM users WHERE id=?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if row and row[0] else None

def send_email_to_user(user_id, subject, body_text):
    """向指定用户邮箱发送邮件。若 SMTP 未配置或用户无邮箱则静默跳过。"""
    if not _smtp_configured():
        return False, 'SMTP 未配置'
    to_email = _get_user_email(user_id)
    if not to_email or '@' not in to_email:
        return False, '用户未设置有效邮箱'
    c = _get_smtp_config()
    host, port, user, password, from_addr = c['host'], c['port'], c['user'], c['password'], c['from_addr'] or c['user']
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP(host, port) as server:
            server.starttls(context=context)
            if password:
                server.login(user, password)
            msg = f"Subject: {subject}\nTo: {to_email}\nFrom: {from_addr}\nContent-Type: text/plain; charset=utf-8\n\n{body_text}"
            server.sendmail(from_addr, [to_email], msg.encode('utf-8'))
        return True, None
    except Exception as e:
        return False, str(e)

# CORS：生产环境通过 CORS_ORIGINS 配置允许的前端域名（逗号分隔），未设置则允许所有（开发用）
_cors_origins = os.environ.get('CORS_ORIGINS', '').strip()
if _cors_origins:
    _origins_list = [o.strip() for o in _cors_origins.split(',') if o.strip()]
else:
    _origins_list = ["*"]
CORS(app, resources={
    r"/api/*": {
        "origins": _origins_list,
        "methods": ["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "Access-Control-Allow-Origin", "X-Requested-With"],
        "supports_credentials": True
    }
}, supports_credentials=True)


@app.errorhandler(500)
def internal_error(e):
    logger.exception('Internal server error')
    # 静态资源或打包运行时返回真实错误便于排查白屏
    if request.path.startswith('/static/') or getattr(sys, 'frozen', False):
        return jsonify({'message': str(e), 'path': request.path}), 500
    if _is_production:
        return jsonify({'message': '服务器内部错误，请稍后再试'}), 500
    return jsonify({'message': str(e)}), 500


@app.errorhandler(Exception)
def handle_exception(e):
    if isinstance(e, (AttributeError, ValueError, KeyError)):
        logger.warning('Bad request: %s', e)
        return jsonify({'message': '请求参数错误' if _is_production else str(e)}), 400
    logger.exception('Unhandled exception')
    if request.path.startswith('/static/') or getattr(sys, 'frozen', False):
        return jsonify({'message': str(e), 'path': request.path}), 500
    if _is_production:
        return jsonify({'message': '服务器内部错误，请稍后再试'}), 500
    return jsonify({'message': str(e)}), 500


# 简单限流：按 IP+端点 每分钟最多 N 次（用于登录/注册/忘记密码）
_rate_limit_store = {}
_RATE_LIMIT_WINDOW = 60  # 秒
_RATE_LIMIT_MAX = 6      # 每窗口最多次数

def _rate_limit_key():
    return (request.remote_addr or '127.0.0.1', request.path)

def _check_rate_limit():
    now = time.time()
    key = _rate_limit_key()
    if key not in _rate_limit_store:
        _rate_limit_store[key] = []
    times = _rate_limit_store[key]
    times[:] = [t for t in times if now - t < _RATE_LIMIT_WINDOW]
    if len(times) >= _RATE_LIMIT_MAX:
        return False
    times.append(now)
    return True


@app.route('/api/health', methods=['GET', 'HEAD'])
def health():
    """健康检查，供负载均衡或监控探测"""
    try:
        db_path = _DB_PATH
        if os.path.isfile(db_path):
            conn = sqlite3.connect(db_path)
            conn.cursor().execute('SELECT 1').fetchone()
            conn.close()
        return '', 200
    except Exception as e:
        logger.warning('Health check failed: %s', e)
        return jsonify({'status': 'unhealthy', 'message': str(e) if not _is_production else 'db_error'}), 503


# Database initialization
# 生产环境请设置 SKIP_DROP_TABLES=1 或 FLASK_ENV=production，仅执行迁移（加表/加列），不 DROP 表
def init_db():
    conn = sqlite3.connect(_DB_PATH, timeout=30)
    cursor = conn.cursor()
    cursor.execute("PRAGMA journal_mode=WAL")  # 允许读与写并发，减轻 database is locked
    skip_drop = os.environ.get('SKIP_DROP_TABLES', '').strip() in ('1', 'true', 'yes') or os.environ.get('FLASK_ENV') == 'production'

    if not skip_drop:
        # 开发环境：先删依赖 companies 的表，再删 companies，避免重启后「最新企业资讯」仍显示已删企业
        for t in ('risk_alerts', 'company_media_reviews', 'company_news', 'company_keywords', 'user_company_favorites', 'user_companies'):
            cursor.execute("DROP TABLE IF EXISTS " + t)
        cursor.execute("DROP TABLE IF EXISTS risk_insights")
        cursor.execute("DROP TABLE IF EXISTS news_items")
        cursor.execute("DROP TABLE IF EXISTS documents")
        cursor.execute("DROP TABLE IF EXISTS risk_alerts")
        cursor.execute("DROP TABLE IF EXISTS companies")
        cursor.execute("DROP TABLE IF EXISTS users")

    # 生产环境（skip_drop）也需确保 users/companies 等核心表存在（新建数据目录时）
    if skip_drop:
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            industry TEXT,
            risk_level TEXT DEFAULT '未知',
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            legal_representative TEXT,
            registered_capital TEXT,
            business_status TEXT,
            registered_address TEXT,
            business_scope TEXT,
            equity_structure TEXT,
            social_evaluation TEXT,
            crawl_status TEXT DEFAULT 'pending',
            established_date TEXT,
            legal_cases TEXT,
            equity_changes TEXT,
            capital_changes TEXT
        )
        ''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS risk_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER,
            alert_type TEXT NOT NULL,
            severity TEXT NOT NULL,
            description TEXT,
            source TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
        ''')
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO users (username, email, password_hash, role) VALUES (?, ?, ?, ?)",
                ("admin", "admin@example.com", generate_password_hash("admin123"), "admin"))
        conn.commit()
    # 开发环境：创建表并插入示例数据
    if not skip_drop:
        # Users table
        cursor.execute('''
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        # Companies table
        cursor.execute('''
        CREATE TABLE companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            industry TEXT,
            risk_level TEXT DEFAULT '未知',
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            legal_representative TEXT,
            registered_capital TEXT,
            business_status TEXT,
            registered_address TEXT,
            business_scope TEXT,
            equity_structure TEXT,
            social_evaluation TEXT,
            crawl_status TEXT DEFAULT 'pending',
            established_date TEXT,
            legal_cases TEXT,
            equity_changes TEXT,
            capital_changes TEXT
        )
        ''')
        # Risk alerts table
        cursor.execute('''
        CREATE TABLE risk_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER,
            alert_type TEXT NOT NULL,
            severity TEXT NOT NULL,
            description TEXT,
            source TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
        ''')
        # Documents table
        cursor.execute('''
        CREATE TABLE documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            company_id INTEGER,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'uploaded',
            analysis_result TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
        ''')
        # News items table
        cursor.execute('''
        CREATE TABLE news_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            category TEXT NOT NULL,
            source_url TEXT,
            scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            user_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
        ''')
        # Risk insights table
        cursor.execute('''
        CREATE TABLE risk_insights (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            content TEXT NOT NULL,
            risk_level TEXT NOT NULL,
            source_data TEXT,
            generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            user_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
        ''')
        # Insert sample data
        cursor.execute("INSERT INTO users (username, email, password_hash, role) VALUES (?, ?, ?, ?)",
                   ("admin", "admin@example.com", generate_password_hash("admin123"), "admin"))
        cursor.execute("INSERT INTO companies (name, industry) VALUES (?, ?)", 
                   ("ABC科技有限公司", "科技"))
        cursor.execute("INSERT INTO companies (name, industry) VALUES (?, ?)",
                   ("XYZ制造集团", "制造业"))
        cursor.execute("INSERT INTO companies (name, industry) VALUES (?, ?)",
                   ("EFG金融服务", "金融"))
        # Insert sample news data
        sample_news = [
        ("某科技公司因数据泄露面临重大法律风险", "根据最新报道，某知名科技公司因数据泄露事件面临重大法律风险。该公司未能及时保护用户隐私数据，导致大量敏感信息外泄。目前监管部门已介入调查，可能面临数百万美元的罚款。此事件提醒所有企业在数字化转型过程中必须加强数据安全防护措施。", "法律", "https://example.com/news/1", 1),
        ("新兴行业监管政策更新，多家企业需调整合规策略", "新兴行业的监管政策迎来重大更新，多家相关企业需紧急调整合规策略。新政策涵盖了数据处理、消费者保护、反垄断等多个方面。专家建议企业应尽快评估新政策对业务的影响，并制定相应的合规计划。预计未来几个月内，行业将迎来一波合规调整潮。", "监管", "https://example.com/news/2", 1),
        ("金融行业信用风险指数上升至警戒水平", "最新发布的金融行业信用风险指数已上升至警戒水平，引发业界关注。受经济环境变化和市场波动影响，多个金融机构的信贷质量出现下滑。分析师指出，银行等金融机构应加强对贷款组合的风险管理，提前做好应对准备。", "金融", "https://example.com/news/3", 1),
        ("供应链风险评估显示多个关键节点存在隐患", "一项全面的供应链风险评估显示，多个关键节点存在潜在隐患。这些隐患主要包括单一供应商依赖、地理集中度高、应急响应能力不足等问题。专家建议企业应建立多元化的供应网络，提高供应链韧性，并制定详细的应急预案。", "运营", "https://example.com/news/4", 1),
        ("国际制裁影响扩大，跨境业务企业面临新挑战", "随着国际制裁范围的进一步扩大，从事跨境业务的企业面临新的挑战。受影响的企业需要重新评估其国际业务布局，调整供应链结构，并确保遵守各司法管辖区的法律法规。企业应密切关注政策变化，及时调整战略方向。", "国际", "https://example.com/news/5", 1)
        ]
        for news in sample_news:
            cursor.execute("INSERT INTO news_items (title, content, category, source_url, user_id) VALUES (?, ?, ?, ?, ?)", news)
        # Insert sample insights data
        sample_insights = [
        ("全球供应链风险", "地缘政治紧张局势影响全球供应链稳定性", "当前全球供应链面临多重挑战，包括地缘政治紧张局势、贸易摩擦、自然灾害等因素。企业应采取多元化供应策略，建立弹性供应链体系，以应对不确定性风险。建议企业加强供应链可视化管理，提高风险预警能力。", "高危", "", 1),
        ("金融监管趋严", "新法规出台，金融机构合规成本上升", "随着金融监管政策的不断收紧，金融机构面临更高的合规要求。新法规涵盖资本充足率、风险管理、信息披露等多个方面。机构需投入更多资源进行合规建设，同时也推动了行业的规范化发展。", "中危", "", 1),
        ("网络安全威胁", "勒索软件攻击频率增加，企业面临数据泄露风险", "网络安全威胁持续升级，特别是勒索软件攻击频率显著增加。企业需要建立全面的安全防护体系，包括网络隔离、数据备份、员工培训等措施。同时，应制定应急响应计划，确保在遭受攻击时能够快速恢复。", "中高危", "", 1)
        ]
        for insight in sample_insights:
            cursor.execute("INSERT INTO risk_insights (title, description, content, risk_level, source_data, user_id) VALUES (?, ?, ?, ?, ?, ?)", insight)
    # 以下为迁移：仅加表/加列，不删数据（生产与开发均执行）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS company_media_reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            platform TEXT,
            title TEXT,
            content TEXT,
            sentiment TEXT,
            source_url TEXT,
            llm_summary TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    # LLM 配置表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS llm_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL UNIQUE,
            api_key TEXT,
            base_url TEXT,
            model TEXT DEFAULT 'gpt-4o-mini',
            enable_web_search INTEGER DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    # 兼容旧表：添加 enable_web_search 列（若不存在）
    try:
        cursor.execute("ALTER TABLE llm_config ADD COLUMN enable_web_search INTEGER DEFAULT 0")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 兼容旧表：documents 添加 analysis_result（若表存在且无此列）
    try:
        cursor.execute("ALTER TABLE documents ADD COLUMN analysis_result TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 兼容旧表：companies 添加 llm_status, media_status, news_status（三灯：工商、相关新闻、社会评价）
    for col in ('llm_status', 'media_status', 'news_status'):
        try:
            cursor.execute(f"ALTER TABLE companies ADD COLUMN {col} TEXT DEFAULT 'pending'")
            conn.commit()
        except sqlite3.OperationalError:
            pass
    # 用户收藏企业表（红心：仅影响是否标星，取消收藏后企业仍在「我的企业」中）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_company_favorites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            company_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, company_id),
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    # 用户关联企业表：用户添加的企业（与收藏分离，取消红心后企业仍可见）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_companies (
            user_id INTEGER NOT NULL,
            company_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, company_id),
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    try:
        cursor.execute("INSERT OR IGNORE INTO user_companies (user_id, company_id) SELECT user_id, company_id FROM user_company_favorites")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 用户提交链接表
    # 企业关键词（词云）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS company_keywords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            keyword TEXT NOT NULL,
            weight INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    # 企业关联新闻/资讯（滚动）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS company_news (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER,
            title TEXT NOT NULL,
            content TEXT,
            source TEXT,
            source_url TEXT,
            sentiment_score REAL,
            risk_level TEXT,
            category TEXT,
            keywords TEXT,
            publish_date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    # 用户设置：搜索间隔、报告格式
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL UNIQUE,
            search_interval_minutes INTEGER DEFAULT 30,
            report_template TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    if not skip_drop:
        cursor.execute("SELECT id, name FROM companies LIMIT 3")
        for cid, cname in cursor.fetchall():
            cursor.execute("""INSERT INTO company_news (company_id, title, content, source, sentiment_score, risk_level, category, publish_date)
                VALUES (?,?,?,?,?,?,?,?)""",
                (cid, f'关于{cname}的最新动态', '企业持续稳健发展，市场关注度较高。', '示例来源', 0.6, '低', '经营', datetime.now().strftime('%Y-%m-%d')))
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            company_id INTEGER,
            url TEXT NOT NULL,
            title TEXT,
            status TEXT DEFAULT 'pending',
            analysis_result TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    # 尽调补充：企业 AI 助手对话记录与归集摘要
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS company_supplements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            merged_fields TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id),
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    try:
        cursor.execute("ALTER TABLE companies ADD COLUMN supplement_notes TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 宏观政策摘要（大模型每日更新，右侧面板展示）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS macro_policy_digest (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 政策与市场环境：多维度新闻，每条单独一条记录（不按时间跨度总结）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS macro_policy_news (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            source TEXT,
            dimension TEXT,
            published_at TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 宏观指数：按日汇总，由 macro_policy_news 计算得到，供风险预测等使用
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS macro_daily_index (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts_date TEXT NOT NULL UNIQUE,
            policy_score REAL NOT NULL DEFAULT 0,
            macro_risk_score REAL NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 企业事件特征表：新闻 / 舆情 / 政策 / 工商变更 等结构化事件，供风险聚合与预测使用
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS enterprise_event_feature (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            enterprise_id INTEGER NOT NULL,
            source_type TEXT NOT NULL,          -- NEWS / SOCIAL / POLICY / BIZ_REG
            source_id INTEGER,                  -- 对应原始记录主键（如 company_news.id）
            event_date TEXT NOT NULL,           -- 事件日期（YYYY-MM-DD）
            event_type TEXT NOT NULL,           -- LEGAL_PENALTY / LITIGATION / FINANCIAL_STRESS / MANAGEMENT_CHANGE / PUBLIC_OPINION / OTHER ...
            sentiment_label TEXT,               -- NEG / NEU / POS
            sentiment_score REAL,               -- -1 ~ 1
            severity_score REAL,                -- 0 ~ 1
            policy_direction TEXT,              -- POSITIVE / NEUTRAL / NEGATIVE（仅 POLICY）
            policy_strength REAL,               -- 0 ~ 1（仅 POLICY）
            biz_change_type TEXT,               -- 高管变更 / 注册资本变更 等（仅 BIZ_REG）
            extra_json TEXT,                    -- 其他 LLM 抽取字段（JSON）
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (enterprise_id) REFERENCES companies (id)
        )
    ''')
    # 企业风险时间序列表：按日聚合各维度风险得分与辅助特征
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS enterprise_risk_timeseries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            enterprise_id INTEGER NOT NULL,
            ts_date TEXT NOT NULL,              -- 日期（YYYY-MM-DD）
            score_legal REAL NOT NULL,          -- 0~1 法律与合规风险
            score_business REAL NOT NULL,       -- 0~1 经营与财务风险
            score_media REAL NOT NULL,          -- 0~1 舆情与媒体风险
            score_policy REAL NOT NULL,         -- 0~1 政策与监管风险
            score_industry REAL NOT NULL,       -- 0~1 行业与宏观风险
            risk_score REAL NOT NULL,           -- 0~100 综合风险得分
            news_count INTEGER NOT NULL DEFAULT 0,
            neg_news_ratio REAL NOT NULL DEFAULT 0,
            sentiment_index REAL,               -- 当日情绪指数（平均 sentiment_score）
            sentiment_vol REAL,                 -- 过去 N 日情绪波动率
            policy_impact REAL,                 -- 政策负面影响指数
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(enterprise_id, ts_date),
            FOREIGN KEY (enterprise_id) REFERENCES companies (id)
        )
    ''')
    # 预测回测指标：MAE、残差标准差等，供预测区间与准确性评估
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS backtest_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            metric_name TEXT NOT NULL UNIQUE,
            value_real REAL,
            value_text TEXT,
            extra_json TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 预测与宏观参数：可配置权重与中性值，便于客观调参
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS prediction_config (
            key TEXT NOT NULL PRIMARY KEY,
            value_text TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    for cfg_key, cfg_val in [('macro_adjustment_scale', '20'), ('macro_neutral', '0.5'), ('prediction_interval_z', '1.96')]:
        try:
            cursor.execute("INSERT OR IGNORE INTO prediction_config (key, value_text, updated_at) VALUES (?, ?, ?)", (cfg_key, cfg_val, datetime.now().isoformat()))
        except sqlite3.OperationalError:
            pass
    # 企业风险标签：用于训练 ML 模型（如爆雷/高风险事件标签）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS enterprise_risk_label (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            enterprise_id INTEGER NOT NULL,
            as_of_date TEXT NOT NULL,           -- 标签对应的观察起点日（YYYY-MM-DD）
            label_type TEXT NOT NULL,           -- explosion / default / penalty 等
            label_value INTEGER NOT NULL,       -- 0/1
            note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(enterprise_id, as_of_date, label_type),
            FOREIGN KEY (enterprise_id) REFERENCES companies (id)
        )
    ''')
    # company_news 多维度风险
    try:
        cursor.execute("ALTER TABLE company_news ADD COLUMN risk_dimensions TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 企业专属标签：写入时存当前企业名，筛选时只返回标签与当前企业名一致的新闻，避免串企业
    try:
        cursor.execute("ALTER TABLE company_news ADD COLUMN company_tag TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 一次性清空 company/4 相关新闻（按用户要求不再显示）
    try:
        cursor.execute("DELETE FROM company_news WHERE company_id=4")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 系统设置（管理员在管理后台配置，如 SMTP 发件人）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    # 操作审计日志（谁在何时做了什么，便于合规）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            resource_type TEXT,
            resource_id TEXT,
            detail TEXT,
            ip TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    # 预警规则（可配置：风险升级、分类阈值、情感阈值等）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alert_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            rule_type TEXT NOT NULL,
            config TEXT,
            enabled INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    try:
        cursor.execute("ALTER TABLE companies ADD COLUMN tags TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 站内通知
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            type TEXT NOT NULL,
            title TEXT NOT NULL,
            body TEXT,
            related_type TEXT,
            related_id TEXT,
            read_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    # 任务执行记录（爬取/新闻搜索等）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS task_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_type TEXT NOT NULL,
            company_id INTEGER,
            status TEXT NOT NULL,
            message TEXT,
            started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            finished_at TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )
    ''')
    try:
        cursor.execute("ALTER TABLE risk_alerts ADD COLUMN processed_at TIMESTAMP")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute("ALTER TABLE company_news ADD COLUMN risk_dimensions TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    # 用户扩展设置与忘记密码
    for col, ctype in [
        ('alert_threshold', 'TEXT'), ('two_factor_enabled', 'INTEGER DEFAULT 0'),
        ('two_factor_secret', 'TEXT'),
        ('backup_enabled', 'INTEGER DEFAULT 0'), ('backup_frequency', 'TEXT'),
        ('reset_token', 'TEXT'), ('reset_token_expires', 'TEXT')
    ]:
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN %s %s" % (col, ctype))
            conn.commit()
        except sqlite3.OperationalError:
            pass
    
    conn.commit()
    conn.close()

# JWT token decorator
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            logger.info('Auth failed: missing token from %s', request.remote_addr)
            return jsonify({'message': 'Token is missing'}), 401
        
        try:
            token = token.split(" ")[1]  # Remove "Bearer " prefix
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            current_user_id = data['user_id']
        except jwt.ExpiredSignatureError:
            logger.info('Auth failed: token expired from %s', request.remote_addr)
            return jsonify({'message': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            logger.info('Auth failed: invalid token from %s', request.remote_addr)
            return jsonify({'message': 'Token is invalid'}), 401
        
        return f(current_user_id, *args, **kwargs)
    
    return decorated


def admin_required(f):
    """要求当前用户为 admin 角色，与 token_required 同时使用：先 token_required 再 admin_required"""
    @wraps(f)
    def decorated(current_user_id, *args, **kwargs):
        conn = sqlite3.connect(_DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT role FROM users WHERE id = ?", (current_user_id,))
        row = cur.fetchone()
        conn.close()
        if not row or row[0] != 'admin':
            logger.info('Admin required: user_id=%s denied', current_user_id)
            return jsonify({'message': '需要管理员权限'}), 403
        return f(current_user_id, *args, **kwargs)
    return decorated


def _user_company_ids(cursor, user_id):
    """返回当前用户关联且仍存在的企业 id 列表（user_companies），用于数据隔离。"""
    cursor.execute("""SELECT u.company_id FROM user_companies u
                      INNER JOIN companies c ON c.id = u.company_id
                      WHERE u.user_id = ?""", (user_id,))
    return [row[0] for row in cursor.fetchall()]


def _audit_log(user_id, action, resource_type=None, resource_id=None, detail=None, cursor_=None, conn=None):
    """写入操作审计日志。可传入已有 cursor/conn，否则新建连接。"""
    ip = request.remote_addr if request else None
    if cursor_ and conn:
        try:
            cursor_.execute("""INSERT INTO audit_log (user_id, action, resource_type, resource_id, detail, ip)
                VALUES (?,?,?,?,?,?)""", (user_id, action, resource_type or '', str(resource_id) if resource_id is not None else '', detail or '', ip))
            conn.commit()
        except Exception:
            pass
        return
    try:
        c = _db_conn()
        cur = c.cursor()
        cur.execute("""INSERT INTO audit_log (user_id, action, resource_type, resource_id, detail, ip)
            VALUES (?,?,?,?,?,?)""", (user_id, action, resource_type or '', str(resource_id) if resource_id is not None else '', detail or '', ip))
        c.commit()
        c.close()
    except Exception:
        pass


# 处理预检请求
@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = jsonify()
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "*")
        response.headers.add('Access-Control-Allow-Methods', "*")
        return response

# 添加跨域头
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization,X-Requested-With')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

# Authentication routes
@app.route('/api/auth/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return jsonify({})
    if not _check_rate_limit():
        logger.warning('Rate limit exceeded: login from %s', request.remote_addr)
        return jsonify({'message': '请求过于频繁，请稍后再试'}), 429
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, username, password_hash, role, two_factor_enabled FROM users WHERE username=?",
        (username,)
    )
    user = cursor.fetchone()
    conn.close()
    
    if not user or not check_password_hash(user[2], password):
        logger.info('Login failed: username=%s ip=%s', (username or '')[:32], request.remote_addr)
        return jsonify({'message': 'Invalid credentials'}), 401

    user_id, uname, _, role = user[0], user[1], user[2], user[3]
    two_factor_enabled = bool(user[4]) if len(user) > 4 and user[4] is not None else False

    if two_factor_enabled:
        temp_token = jwt.encode({
            'user_id': user_id,
            'username': uname,
            'role': role,
            'type': '2fa_pending',
            'exp': datetime.utcnow() + timedelta(minutes=5)
        }, app.config['SECRET_KEY'], algorithm="HS256")
        return jsonify({
            'requires_2fa': True,
            'temp_token': temp_token,
            'message': '请输入两步验证码'
        }), 200

    token = jwt.encode({
        'user_id': user_id,
        'username': uname,
        'role': role,
        'exp': datetime.utcnow() + timedelta(hours=24)
    }, app.config['SECRET_KEY'], algorithm="HS256")
    _audit_log(user_id, 'login', 'user', user_id, '登录成功')
    return jsonify({
        'token': token,
        'user': {'id': user_id, 'username': uname, 'role': role}
    })


def _temp_2fa_token_required(f):
    """仅接受 2fa_pending 临时 token，用于登录第二步验证。"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': '缺少临时验证令牌'}), 401
        try:
            token = token.split(" ")[1]
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            if data.get('type') != '2fa_pending':
                return jsonify({'message': '无效的临时令牌'}), 401
            current_user_id = data['user_id']
        except jwt.ExpiredSignatureError:
            return jsonify({'message': '验证已过期，请重新登录'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'message': '无效的临时令牌'}), 401
        return f(current_user_id, *args, **kwargs)
    return decorated


@app.route('/api/auth/2fa/setup', methods=['POST', 'OPTIONS'])
@token_required
def two_fa_setup(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': '用户不存在'}), 404
    username = row[0]
    secret = pyotp.random_base32()
    cursor.execute("UPDATE users SET two_factor_secret=?, two_factor_enabled=0 WHERE id=?", (secret, current_user_id))
    conn.commit()
    conn.close()
    totp = pyotp.TOTP(secret)
    otpauth_url = totp.provisioning_uri(name=username, issuer_name='风险洞察平台')
    img = qrcode.make(otpauth_url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    return jsonify({
        'secret': secret,
        'otpauth_url': otpauth_url,
        'qr_image': 'data:image/png;base64,' + qr_base64
    })


@app.route('/api/auth/2fa/verify', methods=['POST', 'OPTIONS'])
@token_required
def two_fa_verify(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    code = (data.get('code') or '').strip().replace(' ', '')
    if not code or len(code) != 6:
        return jsonify({'message': '请输入 6 位验证码'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT two_factor_secret FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    conn.close()
    if not row or not row[0]:
        return jsonify({'message': '请先进行两步验证设置'}), 400
    totp = pyotp.TOTP(row[0])
    if not totp.verify(code, valid_window=1):
        return jsonify({'message': '验证码错误或已过期'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET two_factor_enabled=1 WHERE id=?", (current_user_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '两步验证已启用'})


@app.route('/api/auth/2fa/disable', methods=['POST', 'OPTIONS'])
@token_required
def two_fa_disable(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    code = (data.get('code') or '').strip().replace(' ', '')
    if not code or len(code) != 6:
        return jsonify({'message': '请输入 6 位验证码'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT two_factor_secret FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    conn.close()
    if not row or not row[0]:
        conn2 = sqlite3.connect(_DB_PATH)
        cur = conn2.cursor()
        cur.execute("UPDATE users SET two_factor_enabled=0, two_factor_secret=NULL WHERE id=?", (current_user_id,))
        conn2.commit()
        conn2.close()
        _audit_log(current_user_id, '2fa_disable', 'user', current_user_id, '两步验证已关闭')
        return jsonify({'message': '两步验证已关闭'})
    totp = pyotp.TOTP(row[0])
    if not totp.verify(code, valid_window=1):
        return jsonify({'message': '验证码错误或已过期'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET two_factor_enabled=0, two_factor_secret=NULL WHERE id=?", (current_user_id,))
    conn.commit()
    conn.close()
    _audit_log(current_user_id, '2fa_disable', 'user', current_user_id, '两步验证已关闭')
    return jsonify({'message': '两步验证已关闭'})


@app.route('/api/auth/2fa/verify-login', methods=['POST', 'OPTIONS'])
@_temp_2fa_token_required
def two_fa_verify_login(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    code = (data.get('code') or '').strip().replace(' ', '')
    if not code or len(code) != 6:
        return jsonify({'message': '请输入 6 位验证码'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT username, role, two_factor_secret FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    conn.close()
    if not row or not row[2]:
        return jsonify({'message': '两步验证未配置'}), 400
    totp = pyotp.TOTP(row[2])
    if not totp.verify(code, valid_window=1):
        return jsonify({'message': '验证码错误或已过期'}), 400
    token = jwt.encode({
        'user_id': current_user_id,
        'username': row[0],
        'role': row[1],
        'exp': datetime.utcnow() + timedelta(hours=24)
    }, app.config['SECRET_KEY'], algorithm="HS256")
    _audit_log(current_user_id, 'login', 'user', current_user_id, '2FA验证成功')
    return jsonify({
        'token': token,
        'user': {'id': current_user_id, 'username': row[0], 'role': row[1]}
    })


@app.route('/api/auth/me', methods=['GET', 'OPTIONS'])
@token_required
def auth_me(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, email, role, created_at FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'User not found'}), 404
    out = {'id': row[0], 'username': row[1], 'email': row[2], 'role': row[3], 'created_at': row[4]}
    try:
        cursor.execute(
            "SELECT alert_threshold, two_factor_enabled, backup_enabled, backup_frequency FROM users WHERE id=?",
            (current_user_id,)
        )
        r = cursor.fetchone()
        if r:
            out['alert_threshold'] = r[0]
            out['two_factor_enabled'] = bool(r[1]) if r[1] is not None else False
            out['backup_enabled'] = bool(r[2]) if r[2] is not None else False
            out['backup_frequency'] = r[3]
    except sqlite3.OperationalError:
        pass
    conn.close()
    return jsonify(out)

@app.route('/api/auth/change-password', methods=['POST', 'OPTIONS'])
@token_required
def change_password(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    old_pwd = data.get('current_password')
    new_pwd = data.get('new_password')
    if not old_pwd or not new_pwd:
        return jsonify({'detail': '缺少参数'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT password_hash FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    if not row or not check_password_hash(row[0], old_pwd):
        conn.close()
        return jsonify({'detail': '当前密码错误'}), 400
    cursor.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pwd), current_user_id))
    conn.commit()
    conn.close()
    return jsonify({'message': '密码已修改'})

@app.route('/api/auth/me', methods=['PATCH', 'PUT', 'OPTIONS'])
@token_required
def update_profile(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    username = data.get('username')
    email = data.get('email')
    alert_threshold = data.get('alert_threshold')
    backup_enabled = data.get('backup_enabled')
    backup_frequency = data.get('backup_frequency')
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    try:
        if username is not None:
            cursor.execute("UPDATE users SET username=? WHERE id=?", (username.strip(), current_user_id))
        if email is not None:
            cursor.execute("UPDATE users SET email=? WHERE id=?", (email.strip(), current_user_id))
        if alert_threshold is not None:
            try:
                cursor.execute("UPDATE users SET alert_threshold=? WHERE id=?", (str(alert_threshold), current_user_id))
            except sqlite3.OperationalError:
                pass
        if backup_enabled is not None:
            try:
                cursor.execute("UPDATE users SET backup_enabled=? WHERE id=?", (1 if backup_enabled else 0, current_user_id))
            except sqlite3.OperationalError:
                pass
        if backup_frequency is not None:
            try:
                cursor.execute("UPDATE users SET backup_frequency=? WHERE id=?", (str(backup_frequency), current_user_id))
            except sqlite3.OperationalError:
                pass
        conn.commit()
        cursor.execute("SELECT id, username, email, role, created_at FROM users WHERE id=?", (current_user_id,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return jsonify({'message': 'User not found'}), 404
        return jsonify({
            'id': row[0], 'username': row[1], 'email': row[2], 'role': row[3], 'created_at': row[4]
        })
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'message': '用户名或邮箱已被使用'}), 400


@app.route('/api/backup/status', methods=['GET', 'OPTIONS'])
@token_required
def backup_status(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM system_settings WHERE key=?", ('last_backup',))
    row = cursor.fetchone()
    conn.close()
    return jsonify({'last_backup': row[0] if row and row[0] else None})


@app.route('/api/auth/forgot-password', methods=['POST', 'OPTIONS'])
def forgot_password():
    if request.method == 'OPTIONS':
        return jsonify({})
    if not _check_rate_limit():
        logger.warning('Rate limit exceeded: forgot-password from %s', request.remote_addr)
        return jsonify({'message': '请求过于频繁，请稍后再试'}), 429
    data = request.get_json() or {}
    email = (data.get('email') or '').strip()
    if not email:
        return jsonify({'message': '请提供邮箱'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE email=?", (email,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': '若该邮箱已注册，将收到重置链接'}), 200
    reset_token = secrets.token_urlsafe(32)
    expires = (datetime.utcnow() + timedelta(hours=1)).isoformat()
    try:
        cursor.execute(
            "UPDATE users SET reset_token=?, reset_token_expires=? WHERE id=?",
            (reset_token, expires, row[0])
        )
        conn.commit()
    except sqlite3.OperationalError:
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN reset_token TEXT")
            cursor.execute("ALTER TABLE users ADD COLUMN reset_token_expires TEXT")
            cursor.execute(
                "UPDATE users SET reset_token=?, reset_token_expires=? WHERE id=?",
                (reset_token, expires, row[0])
            )
            conn.commit()
        except Exception:
            pass
    conn.close()
    # 若已配置 SMTP：发送重置链接邮件，且不向前端返回 token（安全）；否则开发模式返回 token 便于测试
    if _smtp_configured() and row:
        reset_url = os.environ.get('RESET_PASSWORD_BASE_URL', request.host_url.rstrip('/')) + '/reset-password?token=' + reset_token
        ok, err = send_email_to_user(row[0], '重置密码', '请点击以下链接重置密码（1小时内有效）：\n' + reset_url + '\n\n如非本人操作请忽略。')
        if not ok:
            logger.warning('Forgot password email failed: %s', err)
    out = {'message': '若该邮箱已注册，将收到重置链接'}
    if not _smtp_configured():
        out['reset_token'] = reset_token
    return jsonify(out), 200

@app.route('/api/auth/reset-password', methods=['POST', 'OPTIONS'])
def reset_password():
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    token = (data.get('token') or data.get('reset_token') or '').strip()
    new_password = data.get('new_password') or data.get('password') or ''
    if not token or not new_password:
        return jsonify({'message': '请提供重置令牌和新密码'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, reset_token_expires FROM users WHERE reset_token=?", (token,))
        row = cursor.fetchone()
    except sqlite3.OperationalError:
        conn.close()
        return jsonify({'message': '重置功能不可用'}), 503
    if not row:
        conn.close()
        return jsonify({'message': '无效或过期的重置链接'}), 400
    try:
        expires = datetime.fromisoformat(row[1]) if row[1] else datetime.min
    except Exception:
        expires = datetime.min
    if datetime.utcnow() > expires:
        cursor.execute("UPDATE users SET reset_token=NULL, reset_token_expires=NULL WHERE id=?", (row[0],))
        conn.commit()
        conn.close()
        return jsonify({'message': '重置链接已过期'}), 400
    password_hash = generate_password_hash(new_password)
    cursor.execute("UPDATE users SET password_hash=?, reset_token=NULL, reset_token_expires=NULL WHERE id=?", (password_hash, row[0]))
    conn.commit()
    conn.close()
    return jsonify({'message': '密码已重置，请使用新密码登录'}), 200

@app.route('/api/auth/register', methods=['POST', 'OPTIONS'])
def register():
    if request.method == 'OPTIONS':
        return jsonify({})
    if not _check_rate_limit():
        logger.warning('Rate limit exceeded: register from %s', request.remote_addr)
        return jsonify({'message': '请求过于频繁，请稍后再试'}), 429
    data = request.get_json()
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')
    
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    
    try:
        password_hash = generate_password_hash(password)
        role = (data.get('role') or 'user').strip().lower()
        if role not in ('user', 'admin'):
            role = 'user'
        cursor.execute("INSERT INTO users (username, email, password_hash, role) VALUES (?, ?, ?, ?)",
                       (username, email, password_hash, role))
        conn.commit()
        conn.close()
        return jsonify({'message': 'User registered successfully'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'message': 'Username or email already exists'}), 400

@app.route('/api/notifications/email-status', methods=['GET', 'OPTIONS'])
@token_required
def email_status(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    configured = _smtp_configured()
    user_email = _get_user_email(current_user_id) if configured else None
    return jsonify({
        'smtp_configured': configured,
        'user_email': user_email or '',
        'hint': '请在个人中心填写邮箱；SMTP 由管理员通过环境变量配置（SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SMTP_FROM）。'
    })

@app.route('/api/notifications/test-email', methods=['POST', 'OPTIONS'])
@token_required
def send_test_email(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    ok, err = send_email_to_user(
        current_user_id,
        '【风险监控平台】邮件通知测试',
        '这是一封测试邮件。若您收到此邮件，说明系统已能将提醒发送至您在个人中心设置的邮箱。'
    )
    if ok:
        return jsonify({'message': '测试邮件已发送，请查收'})
    return jsonify({'message': err or '发送失败'}), 400

@app.route('/api/notifications/send-risk-digest', methods=['POST', 'OPTIONS'])
@token_required
def send_risk_digest(current_user_id):
    """向当前用户邮箱发送风险速览（高风险企业数量等）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT risk_level, COUNT(*) FROM companies GROUP BY risk_level")
    rows = cursor.fetchall()
    conn.close()
    high = medium = low = 0
    for r in rows:
        if r[0] == '高':
            high = r[1]
        elif r[0] == '中':
            medium = r[1]
        else:
            low += r[1]
    body = f"""风险监控平台 - 风险速览

当前监控企业风险分布：
- 高风险：{high} 家
- 中风险：{medium} 家
- 低/未知：{low} 家

请登录系统查看详情。"""
    ok, err = send_email_to_user(current_user_id, '【风险监控平台】风险速览提醒', body)
    if ok:
        return jsonify({'message': '风险速览已发送至您的邮箱'})
    return jsonify({'message': err or '发送失败'}), 400

# News and Insights routes
@app.route('/api/news', methods=['GET', 'OPTIONS'])
@token_required
def get_news(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
        
    category = request.args.get('category')
    skip = int(request.args.get('skip', 0))
    limit = int(request.args.get('limit', 10))
    
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    
    query = "SELECT id, title, content, category, source_url, scraped_at, created_at FROM news_items WHERE user_id = ?"
    params = [current_user_id]
    
    if category:
        query += " AND category = ?"
        params.append(category)
    
    query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    params.extend([limit, skip])
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    news_items = []
    for row in rows:
        news_items.append({
            'id': row[0],
            'title': row[1],
            'content': row[2],
            'category': row[3],
            'source_url': row[4],
            'scraped_at': row[5],
            'created_at': row[6]
        })
    
    conn.close()
    return jsonify(news_items)

@app.route('/api/risk-insights', methods=['GET', 'OPTIONS'])
@token_required
def get_risk_insights(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
        
    risk_level = request.args.get('risk_level')
    skip = int(request.args.get('skip', 0))
    limit = int(request.args.get('limit', 10))
    
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    
    query = "SELECT id, title, description, content, risk_level, generated_at, created_at FROM risk_insights WHERE user_id = ?"
    params = [current_user_id]
    
    if risk_level:
        query += " AND risk_level = ?"
        params.append(risk_level)
    
    query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    params.extend([limit, skip])
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    insights = []
    for row in rows:
        insights.append({
            'id': row[0],
            'title': row[1],
            'description': row[2],
            'content': row[3],
            'risk_level': row[4],
            'generated_at': row[5],
            'created_at': row[6]
        })
    
    conn.close()
    return jsonify(insights)


@app.route('/api/macro-policy-digest', methods=['GET', 'OPTIONS'])
@token_required
def get_macro_policy_digest(current_user_id):
    """获取最新的宏观政策摘要（大模型每日更新）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    db_path = _DB_PATH
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, title, content, updated_at FROM macro_policy_digest ORDER BY updated_at DESC LIMIT 1"
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'title': '', 'content': '', 'updated_at': None})
    return jsonify({
        'id': row[0],
        'title': row[1] or '',
        'content': row[2] or '',
        'updated_at': row[3]
    })


@app.route('/api/macro-policy-digest/refresh', methods=['POST', 'OPTIONS'])
@token_required
def refresh_macro_policy_digest_api(current_user_id):
    """手动触发宏观政策摘要更新"""
    if request.method == 'OPTIONS':
        return jsonify({})
    try:
        from backend.services.scheduler_service import refresh_macro_policy_digest
        ok, msg = refresh_macro_policy_digest()
        if ok:
            return jsonify({'success': True, 'message': msg or '已更新'})
        return jsonify({'success': False, 'message': msg or '更新失败'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/macro-policy-news', methods=['GET', 'OPTIONS'])
@token_required
def get_macro_policy_news(current_user_id):
    """获取政策与市场环境新闻列表（多维度、每条单独，定时自动更新）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    db_path = _DB_PATH
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT id, title, content, source, dimension, published_at, created_at FROM macro_policy_news ORDER BY created_at DESC LIMIT 50"
        )
        rows = cursor.fetchall()
    finally:
        conn.close()
    items = [
        {
            'id': r[0],
            'title': r[1] or '',
            'content': r[2] or '',
            'source': r[3] or '',
            'dimension': r[4] or '政策',
            'published_at': r[5],
            'created_at': r[6],
        }
        for r in rows
    ]
    return jsonify(items)


@app.route('/api/macro-index/refresh', methods=['POST', 'OPTIONS'])
@token_required
def refresh_macro_index_api(current_user_id):
    """根据当前 macro_policy_news 重新计算宏观指数并写入 macro_daily_index（供预测使用）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    try:
        from backend.services import macro_index_service as mis
        mis.DB_PATH = _DB_PATH
        if mis.compute_and_save_macro_index(_DB_PATH):
            return jsonify({'success': True, 'message': '宏观指数已更新'})
        return jsonify({'success': False, 'message': '暂无政策与市场环境数据，请先刷新右侧政策新闻或等待定时更新'}), 400
    except Exception as e:
        logger.exception('refresh_macro_index_api')
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/test-data', methods=['POST', 'OPTIONS'])
@token_required
def add_test_data(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
        
    # Add test news data if not exists
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    
    # Check if test data already exists for this user
    cursor.execute("SELECT COUNT(*) FROM news_items WHERE user_id = ?", (current_user_id,))
    news_count_before = cursor.fetchone()[0]
    
    if news_count_before == 0:
        test_news = [
            ("某科技公司因数据泄露面临重大法律风险", "根据最新报道，某知名科技公司因数据泄露事件面临重大法律风险。该公司未能及时保护用户隐私数据，导致大量敏感信息外泄。目前监管部门已介入调查，可能面临数百万美元的罚款。此事件提醒所有企业在数字化转型过程中必须加强数据安全防护措施。", "法律", "https://example.com/news/1", current_user_id),
            ("新兴行业监管政策更新，多家企业需调整合规策略", "新兴行业的监管政策迎来重大更新，多家相关企业需紧急调整合规策略。新政策涵盖了数据处理、消费者保护、反垄断等多个方面。专家建议企业应尽快评估新政策对业务的影响，并制定相应的合规计划。预计未来几个月内，行业将迎来一波合规调整潮。", "监管", "https://example.com/news/2", current_user_id),
            ("金融行业信用风险指数上升至警戒水平", "最新发布的金融行业信用风险指数已上升至警戒水平，引发业界关注。受经济环境变化和市场波动影响，多个金融机构的信贷质量出现下滑。分析师指出，银行等金融机构应加强对贷款组合的风险管理，提前做好应对准备。", "金融", "https://example.com/news/3", current_user_id),
            ("供应链风险评估显示多个关键节点存在隐患", "一项全面的供应链风险评估显示，多个关键节点存在潜在隐患。这些隐患主要包括单一供应商依赖、地理集中度高、应急响应能力不足等问题。专家建议企业应建立多元化的供应网络，提高供应链韧性，并制定详细的应急预案。", "运营", "https://example.com/news/4", current_user_id),
            ("国际制裁影响扩大，跨境业务企业面临新挑战", "随着国际制裁范围的进一步扩大，从事跨境业务的企业面临新的挑战。受影响的企业需要重新评估其国际业务布局，调整供应链结构，并确保遵守各司法管辖区的法律法规。企业应密切关注政策变化，及时调整战略方向。", "国际", "https://example.com/news/5", current_user_id)
        ]
        
        for news in test_news:
            cursor.execute("INSERT INTO news_items (title, content, category, source_url, user_id) VALUES (?, ?, ?, ?, ?)", news)
    else:
        test_news = []  # Initialize to avoid UnboundLocalError
    
    # Check if test insights already exist for this user
    cursor.execute("SELECT COUNT(*) FROM risk_insights WHERE user_id = ?", (current_user_id,))
    insights_count_before = cursor.fetchone()[0]
    
    if insights_count_before == 0:
        test_insights = [
            ("全球供应链风险", "地缘政治紧张局势影响全球供应链稳定性", "当前全球供应链面临多重挑战，包括地缘政治紧张局势、贸易摩擦、自然灾害等因素。企业应采取多元化供应策略，建立弹性供应链体系，以应对不确定性风险。建议企业加强供应链可视化管理，提高风险预警能力。", "高危", "", current_user_id),
            ("金融监管趋严", "新法规出台，金融机构合规成本上升", "随着金融监管政策的不断收紧，金融机构面临更高的合规要求。新法规涵盖资本充足率、风险管理、信息披露等多个方面。机构需投入更多资源进行合规建设，同时也推动了行业的规范化发展。", "中危", "", current_user_id),
            ("网络安全威胁", "勒索软件攻击频率增加，企业面临数据泄露风险", "网络安全威胁持续升级，特别是勒索软件攻击频率显著增加。企业需要建立全面的安全防护体系，包括网络隔离、数据备份、员工培训等措施。同时，应制定应急响应计划，确保在遭受攻击时能够快速恢复。", "中高危", "", current_user_id)
        ]
        
        for insight in test_insights:
            cursor.execute("INSERT INTO risk_insights (title, description, content, risk_level, source_data, user_id) VALUES (?, ?, ?, ?, ?, ?)", insight)
    else:
        test_insights = []  # Initialize to avoid UnboundLocalError
    
    conn.commit()
    conn.close()
    
    return jsonify({"message": "测试数据已添加", "news_count": len(test_news), "insights_count": len(test_insights)})

# Company routes
@app.route('/api/companies', methods=['GET', 'OPTIONS'])
@token_required
def get_companies(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    search = (request.args.get('search') or '').strip()
    risk_level = (request.args.get('risk_level') or '').strip()
    industry = (request.args.get('industry') or '').strip()
    sort = (request.args.get('sort') or 'last_updated').strip()
    favorite_only = request.args.get('favorite_only', '').lower() in ('1', 'true', 'yes')
    valid_sorts = {'last_updated': 'c.last_updated DESC', 'last_updated_asc': 'c.last_updated ASC',
                  'name': 'c.name ASC', 'name_desc': 'c.name DESC',
                  'risk_level': "CASE WHEN c.risk_level='高' THEN 1 WHEN c.risk_level='中' THEN 2 ELSE 3 END, c.name ASC"}
    order_clause = valid_sorts.get(sort, valid_sorts['last_updated'])

    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT company_id FROM user_company_favorites WHERE user_id=?", (current_user_id,))
    fav_ids = {r[0] for r in cursor.fetchall()}

    # 基础列表：用户关联的企业（user_companies）；favorite_only 时再限定为收藏
    where_parts = ["c.id IN (SELECT company_id FROM user_companies WHERE user_id = ?)"]
    params = [current_user_id]
    if search:
        where_parts.append("(c.name LIKE ? OR c.legal_representative LIKE ?)")
        params.extend(['%' + search + '%', '%' + search + '%'])
    if risk_level:
        where_parts.append("c.risk_level = ?")
        params.append(risk_level)
    if industry:
        where_parts.append("(c.industry LIKE ? OR c.industry = ?)")
        params.extend(['%' + industry + '%', industry])
    if favorite_only:
        where_parts.append("c.id IN (SELECT company_id FROM user_company_favorites WHERE user_id = ?)")
        params.append(current_user_id)  # 仅显示收藏的企业
    where_sql = " AND ".join(where_parts) if where_parts else "1=1"

    try:
        cursor.execute("""
            SELECT c.id, c.name, c.industry, c.risk_level, c.last_updated,
                   COUNT(ra.id) as alert_count, c.legal_representative, c.registered_capital, c.business_status, c.crawl_status,
                   c.llm_status, c.media_status, c.news_status
            FROM companies c
            LEFT JOIN risk_alerts ra ON c.id = ra.company_id
            WHERE """ + where_sql + """
            GROUP BY c.id
            ORDER BY """ + order_clause, tuple(params))
    except sqlite3.OperationalError:
        cursor.execute("""
            SELECT c.id, c.name, c.industry, c.risk_level, c.last_updated,
                   COUNT(ra.id) as alert_count, c.legal_representative, c.registered_capital, c.business_status, c.crawl_status,
                   c.llm_status, c.media_status
            FROM companies c
            LEFT JOIN risk_alerts ra ON c.id = ra.company_id
            WHERE """ + where_sql + """
            GROUP BY c.id
            ORDER BY """ + order_clause, tuple(params))
    company_rows = cursor.fetchall()
    companies = []
    for row in company_rows:
        cid = row[0]
        llm_s = row[10] if len(row) > 10 else 'pending'
        media_s = row[11] if len(row) > 11 else 'pending'
        news_s = row[12] if len(row) > 12 else 'pending'
        companies.append({
            'id': cid,
            'name': row[1],
            'industry': row[2],
            'risk_level': row[3],
            'last_updated': row[4],
            'alert_count': row[5],
            'legal_representative': row[6] or '',
            'registered_capital': row[7] or '',
            'business_status': row[8] or '',
            'crawl_status': row[9] or 'pending',
            'llm_status': llm_s or 'pending',
            'media_status': media_s or 'pending',
            'news_status': news_s or 'pending',
            'is_favorite': cid in fav_ids
        })
    conn.close()
    return jsonify(companies)


@app.route('/api/companies/industries', methods=['GET', 'OPTIONS'])
@token_required
def get_company_industries(current_user_id):
    """返回当前用户企业的行业列表（用于筛选下拉）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify([])
    placeholders = ','.join('?' * len(cids))
    cursor.execute("SELECT DISTINCT industry FROM companies WHERE id IN (" + placeholders + ") AND industry IS NOT NULL AND industry != '' ORDER BY industry", tuple(cids))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([r[0] for r in rows])


@app.route('/api/companies/<int:company_id>', methods=['GET', 'OPTIONS'])
@token_required
def get_company(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
        
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT c.id, c.name, c.industry, c.risk_level, c.last_updated,
                   c.legal_representative, c.registered_capital, c.business_status,
                   c.registered_address, c.business_scope, c.equity_structure,
                   c.social_evaluation, c.crawl_status, c.established_date,
                   c.legal_cases, c.equity_changes, c.capital_changes, c.supplement_notes, c.tags
            FROM companies c
            WHERE c.id = ?
        """, (company_id,))
    except sqlite3.OperationalError:
        cursor.execute("""
            SELECT c.id, c.name, c.industry, c.risk_level, c.last_updated,
                   c.legal_representative, c.registered_capital, c.business_status,
                   c.registered_address, c.business_scope, c.equity_structure,
                   c.social_evaluation, c.crawl_status, c.established_date,
                   c.legal_cases, c.equity_changes, c.capital_changes, c.supplement_notes
            FROM companies c
            WHERE c.id = ?
        """, (company_id,))
    company = cursor.fetchone()
    if not company:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if cursor.fetchone() is None:
        logger.info('Company access denied: user_id=%s company_id=%s', current_user_id, company_id)
        conn.close()
        return jsonify({'message': '无权限访问该企业'}), 403
    
    try:
        cursor.execute("SELECT llm_status, media_status, news_status FROM companies WHERE id=?", (company_id,))
        r = cursor.fetchone()
        if r and len(r) >= 3:
            llm_s, media_s, news_s = (r[0] or 'pending', r[1] or 'pending', r[2] or 'pending')
        elif r:
            llm_s, media_s = (r[0] or 'pending', r[1] or 'pending')
            news_s = 'pending'
        else:
            llm_s = media_s = news_s = 'pending'
    except sqlite3.OperationalError:
        try:
            cursor.execute("SELECT llm_status, media_status FROM companies WHERE id=?", (company_id,))
            r = cursor.fetchone()
            llm_s, media_s = (r[0] or 'pending', r[1] or 'pending') if r else ('pending', 'pending')
            news_s = 'pending'
        except Exception:
            llm_s = media_s = news_s = 'pending'
    except Exception:
        llm_s = media_s = news_s = 'pending'
    cursor.execute("SELECT 1 FROM user_company_favorites WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    is_fav = cursor.fetchone() is not None
    n = len(company)
    result = {
        'id': company[0],
        'name': company[1],
        'industry': company[2],
        'risk_level': company[3],
        'last_updated': company[4],
        'legal_representative': company[5] or '',
        'registered_capital': company[6] or '',
        'business_status': company[7] or '',
        'registered_address': company[8] or '',
        'business_scope': company[9] or '',
        'equity_structure': company[10] or '',
        'social_evaluation': company[11] or '',
        'crawl_status': company[12] or 'pending',
        'llm_status': llm_s,
        'media_status': media_s,
        'news_status': news_s,
        'established_date': company[13] or '',
        'legal_cases': company[14] if n > 14 else '',
        'equity_changes': company[15] if n > 15 else '',
        'capital_changes': company[16] if n > 16 else '',
        'supplement_notes': company[17] if n > 17 else '',
        'tags': company[18] if n > 18 else ''
    }
    
    # Get risk alerts for this company
    cursor.execute("""
        SELECT id, alert_type, severity, description, source, timestamp
        FROM risk_alerts
        WHERE company_id = ?
        ORDER BY timestamp DESC
    """, (company_id,))
    
    alerts = []
    for row in cursor.fetchall():
        alerts.append({
            'id': row[0],
            'alert_type': row[1],
            'severity': row[2],
            'description': row[3],
            'source': row[4],
            'timestamp': row[5]
        })
    
    result['alerts'] = alerts
    cursor.execute("SELECT id, platform, title, content, sentiment, source_url, llm_summary, created_at FROM company_media_reviews WHERE company_id=? ORDER BY created_at DESC", (company_id,))
    result['media_reviews'] = [{'id': r[0], 'platform': r[1], 'title': r[2], 'content': r[3], 'sentiment': r[4], 'source_url': r[5], 'llm_summary': r[6], 'created_at': r[7]} for r in cursor.fetchall()]
    cursor.execute("SELECT keyword, weight FROM company_keywords WHERE company_id=?", (company_id,))
    result['keywords'] = [{'word': r[0], 'weight': r[1]} for r in cursor.fetchall()]
    conn.close()
    
    return jsonify(result)


@app.route('/api/companies/<int:company_id>/risk-timeseries', methods=['GET', 'OPTIONS'])
@token_required
def get_company_risk_timeseries(current_user_id, company_id):
    """返回指定企业近期风险时间序列（供前端画历史风险趋势图）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    days = request.args.get('days', '').strip()
    try:
        days_int = int(days) if days else 90
    except ValueError:
        days_int = 90
    days_int = max(7, min(365, days_int))

    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    # 权限检查：只允许访问自己关联的企业
    cids = _user_company_ids(cursor, current_user_id)
    if company_id not in cids:
        conn.close()
        return jsonify({'message': '无权限访问该企业'}), 403

    cursor.execute(
        """
        SELECT ts_date, score_legal, score_business, score_media,
               score_policy, score_industry, risk_score,
               news_count, neg_news_ratio, sentiment_index, sentiment_vol, policy_impact
        FROM enterprise_risk_timeseries
        WHERE enterprise_id=?
        ORDER BY ts_date DESC
        LIMIT ?
        """,
        (company_id, days_int),
    )
    rows = cursor.fetchall()
    conn.close()
    # 前端画图更习惯时间正序，这里反转
    rows = rows[::-1]
    out = []
    for r in rows:
        out.append({
            'date': r[0],
            'score_legal': r[1],
            'score_business': r[2],
            'score_media': r[3],
            'score_policy': r[4],
            'score_industry': r[5],
            'risk_score': r[6],
            'news_count': r[7],
            'neg_news_ratio': r[8],
            'sentiment_index': r[9],
            'sentiment_vol': r[10],
            'policy_impact': r[11],
        })
    return jsonify(out)


@app.route('/api/v1/predict/risk', methods=['POST', 'OPTIONS'])
@token_required
def predict_risk(current_user_id):
    """
    风险预测接口（阶段 1：简单时间序列外推）。
    请求体：
    {
      "enterprise_id": 1,
      "horizon_days": 30
    }
    """
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    enterprise_id = data.get('enterprise_id')
    if not enterprise_id:
        return jsonify({'message': 'enterprise_id 必填'}), 400
    try:
        enterprise_id = int(enterprise_id)
    except (TypeError, ValueError):
        return jsonify({'message': 'enterprise_id 非法'}), 400

    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if enterprise_id not in cids:
        conn.close()
        return jsonify({'message': '无权限访问该企业'}), 403
    conn.close()

    horizon_days = data.get('horizon_days') or 30
    try:
        horizon_days = int(horizon_days)
    except (TypeError, ValueError):
        horizon_days = 30
    horizon_days = max(7, min(90, horizon_days))

    try:
        from backend.services import prediction_service as ps
        ps.DB_PATH = _DB_PATH
        result = ps.predict_risk_for_enterprise(enterprise_id, horizon_days=horizon_days)
        return jsonify(result)
    except Exception as e:
        logger.exception('predict_risk error')
        return jsonify({'message': '预测失败', 'error': str(e)}), 500


@app.route('/api/v1/predict/backtest', methods=['GET', 'POST', 'OPTIONS'])
@token_required
def predict_backtest(current_user_id):
    """
    运行预测回测，得到 MAE、RMSE、方向准确率、相对朴素基准提升、残差标准差等，
    用于评估准确性并为预测区间提供依据。结果会写入 backtest_metrics 表。
    """
    if request.method == 'OPTIONS':
        return jsonify({})
    lookback = request.args.get('lookback_days', 30)
    horizon = request.args.get('horizon_days', 7)
    try:
        lookback = int(lookback)
        horizon = int(horizon)
    except (TypeError, ValueError):
        lookback, horizon = 30, 7
    lookback = max(7, min(90, lookback))
    horizon = max(1, min(90, horizon))
    try:
        from backend.services import backtest_service as bts
        bts.DB_PATH = _DB_PATH
        result = bts.run_backtest(lookback_days=lookback, horizon_days=horizon, db_path=_DB_PATH)
        return jsonify(result)
    except Exception as e:
        logger.exception('predict_backtest error')
        return jsonify({'error': str(e)}), 500


def _run_company_crawl_and_news_pipeline(company_id, company_name, current_user_id):
    """单企业全流程：工商信息 + 相关新闻搜索 + 媒体舆情爬取。供添加企业、批量导入、重新爬取共用。"""
    try:
        from backend.services.crawler import fetch_company_info, trigger_media_crawl
        from backend.services.llm_service import fetch_company_info_by_llm
        name = company_name
        c0 = _db_conn()
        cur0 = c0.cursor()
        cur0.execute("UPDATE companies SET llm_status='running', news_status='pending', media_status='pending' WHERE id=?", (company_id,))
        cur0.execute("SELECT api_key, base_url, model, enable_web_search FROM llm_config WHERE user_id=?", (current_user_id,))
        llm_row = cur0.fetchone()
        api_key = base_url = model = None
        enable_web_search = False
        if llm_row:
            api_key, base_url, model = llm_row[0], llm_row[1], llm_row[2]
            enable_web_search = bool(llm_row[3]) if len(llm_row) > 3 else False
        c0.commit()
        c0.close()
        if not api_key or not str(api_key).strip():
            print('[工商信息] 未配置 LLM API，使用 enterprise_crawler 模拟数据')
        else:
            print('[工商信息] 使用 LLM 联网搜索，base_url=%s model=%s enable_web=%s' % (base_url or '(默认)', model, enable_web_search))

        info = None
        llm_ok = False
        if api_key:
            llm_info = fetch_company_info_by_llm(name, api_key, base_url, model, enable_web_search=enable_web_search)
            if llm_info:
                print('[工商信息] LLM 返回字段:', list(llm_info.keys()) if llm_info else 'None')
                llm_ok = True
                info = {}
                for k, v in llm_info.items():
                    if v is None:
                        continue
                    if isinstance(v, (list, dict)):
                        info[k] = json.dumps(v, ensure_ascii=False) if v else ''
                    elif v not in ('-', '', '无', '未知'):
                        info[k] = str(v).strip()
                    else:
                        info.setdefault(k, '')
                for f in ('legal_representative','registered_capital','registered_address','industry','business_status','business_scope','equity_structure','established_date','legal_cases','equity_changes','capital_changes'):
                    info.setdefault(f, '')
        if info is None:
            info = fetch_company_info(name)
            print('[工商信息] 使用 enterprise_crawler 回退，info keys:', list(info.keys()) if info else 'None')

        c1 = _db_conn()
        cur1 = c1.cursor()
        cur1.execute("""UPDATE companies SET legal_representative=?, registered_capital=?, business_status=?,
            registered_address=?, business_scope=?, equity_structure=?, established_date=?,
            legal_cases=?, equity_changes=?, capital_changes=?,
            industry=COALESCE(NULLIF(industry,''),?), crawl_status='crawled', llm_status=?, last_updated=? WHERE id=?""",
            (info.get('legal_representative',''), info.get('registered_capital',''), info.get('business_status',''),
             info.get('registered_address',''), info.get('business_scope',''), info.get('equity_structure',''),
             info.get('established_date',''), info.get('legal_cases',''), info.get('equity_changes',''), info.get('capital_changes',''),
             info.get('industry',''), 'success' if llm_ok or info else 'error', datetime.now().isoformat(), company_id))
        c1.commit()
        c1.close()

        c2a = _db_conn()
        c2a.cursor().execute("UPDATE companies SET news_status='running' WHERE id=?", (company_id,))
        c2a.commit()
        c2a.close()
        news_list = None
        try:
            if api_key and str(api_key).strip():
                from backend.services.llm_service import search_company_news
                news_list = search_company_news(name, api_key, base_url or '', model or 'gpt-4o-mini', enable_web_search=enable_web_search)
        except Exception as ne:
            print('[相关新闻] 搜索异常:', ne)
        if news_list is not None:
            c2b = _db_conn()
            cur2 = c2b.cursor()
            cur2.execute("DELETE FROM company_news WHERE company_id=?", (company_id,))
            company_name_tag = (name or '').strip()
            for n in (news_list or []):
                rd = n.get('risk_dimensions') or {}
                rl = n.get('risk_level', '中')
                cur2.execute("""INSERT INTO company_news (company_id, company_tag, title, content, source, sentiment_score, risk_level, category, publish_date, risk_dimensions)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (company_id, company_name_tag, n.get('title', ''), n.get('summary', ''), n.get('source', ''),
                     0.5, rl, n.get('category', '其他'), n.get('date', ''),
                     json.dumps(rd, ensure_ascii=False)))
                if (rl or '').strip() == '高':
                    _create_risk_alert(cur2, company_id, 'news', 'high', (n.get('title') or '')[:200], n.get('source') or '')
                    _notify_alert_created(c2b, company_id, 'news', 'high', (n.get('title') or '')[:200], company_name_tag)
            cur2.execute("UPDATE companies SET news_status=? WHERE id=?", ('success', company_id))
            c2b.commit()
            c2b.close()
            _sync_risk_pipeline_after_news(company_id)
            print('[相关新闻] 已写入 %s 条' % len(news_list or []))
        else:
            c2c = _db_conn()
            c2c.cursor().execute("UPDATE companies SET news_status=? WHERE id=?", ('error', company_id))
            c2c.commit()
            c2c.close()

        c3 = _db_conn()
        cur3 = c3.cursor()
        cur3.execute("UPDATE companies SET social_evaluation=NULL WHERE id=?", (company_id,))
        cur3.execute("DELETE FROM company_media_reviews WHERE company_id=?", (company_id,))
        cur3.execute("UPDATE companies SET media_status='running' WHERE id=?", (company_id,))
        c3.commit()
        c3.close()

        def _save_reviews(cid, reviews):
            from backend.services.media_review_store import save_media_reviews_dedup
            save_media_reviews_dedup(_DB_PATH, cid, name, reviews, api_key, base_url, model)
        def _on_media_error(cid):
            cx = _db_conn()
            cx.cursor().execute("UPDATE companies SET media_status='error' WHERE id=?", (cid,))
            cx.commit()
            cx.close()
        trigger_media_crawl(company_id, name, callback=_save_reviews, on_error=_on_media_error)
    except Exception as e:
        print('Crawl error:', e)
        try:
            cx = _db_conn()
            cu = cx.cursor()
            for col, val in [('llm_status', 'error'), ('media_status', 'error'), ('news_status', 'error')]:
                try:
                    cu.execute("UPDATE companies SET %s=? WHERE id=?" % col, (val, company_id))
                except sqlite3.OperationalError:
                    pass
            cx.commit()
            cx.close()
        except Exception:
            pass


@app.route('/api/companies', methods=['POST', 'OPTIONS'])
@token_required
def add_company(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
        
    data = request.get_json() or {}
    name = data.get('name')
    industry = data.get('industry', '')
    
    if not name:
        return jsonify({'message': 'Company name is required'}), 400
    
    conn = _db_conn()
    cursor = conn.cursor()
    
    try:
        cursor.execute(
            "INSERT INTO companies (name, industry, crawl_status) VALUES (?, ?, 'pending')",
            (name, industry)
        )
        company_id = cursor.lastrowid
        cursor.execute("INSERT OR IGNORE INTO user_companies (user_id, company_id) VALUES (?, ?)", (current_user_id, company_id))
        conn.commit()
        conn.close()

        t = threading.Thread(target=_run_company_crawl_and_news_pipeline, args=(company_id, name, current_user_id))
        t.daemon = True
        t.start()

        return jsonify({
            'id': company_id,
            'name': name,
            'industry': industry,
            'risk_level': '未知',
            'last_updated': datetime.now().isoformat(),
            'crawl_status': 'pending',
            'message': '企业已添加，后台正在自动爬取工商信息与媒体舆情'
        }), 201
    except Exception as e:
        conn.close()
        return jsonify({'message': str(e)}), 500

@app.route('/api/companies/<int:company_id>', methods=['PUT', 'OPTIONS'])
@token_required
def update_company(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        logger.info('Company update denied: user_id=%s company_id=%s', current_user_id, company_id)
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("SELECT id FROM companies WHERE id=?", (company_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cols = []
    vals = []
    for k in ['name','industry','risk_level','legal_representative','registered_capital','business_status','registered_address','business_scope','equity_structure','social_evaluation','established_date','tags','supplement_notes']:
        if k in data:
            cols.append(k + '=?')
            vals.append(data[k])
    if cols:
        old_level = None
        if 'risk_level' in data and (data.get('risk_level') or '').strip() == '高':
            cursor.execute("SELECT risk_level FROM companies WHERE id=?", (company_id,))
            old_row = cursor.fetchone()
            old_level = (old_row[0] or '').strip() if old_row else ''
        cursor.execute("UPDATE companies SET " + ', '.join(cols) + ", last_updated=? WHERE id=?", vals + [datetime.now().isoformat(), company_id])
        if old_level is not None and old_level != '高':
            _create_risk_alert(cursor, company_id, 'company', 'high', '企业风险等级调整为高风险', 'user_edit')
            _notify_alert_created(conn, company_id, 'company', 'high', '企业风险等级调整为高风险')
    conn.commit()
    _audit_log(current_user_id, 'update_company', 'company', company_id, None, cursor_=cursor, conn=conn)
    conn.close()
    return jsonify({'message': '更新成功'})

@app.route('/api/companies/<int:company_id>/favorite', methods=['POST', 'OPTIONS'])
@token_required
def toggle_company_favorite(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM companies WHERE id=?", (company_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("SELECT id FROM user_company_favorites WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    exists = cursor.fetchone()
    if exists:
        cursor.execute("DELETE FROM user_company_favorites WHERE user_id=? AND company_id=?", (current_user_id, company_id))
        conn.commit()
        conn.close()
        return jsonify({'message': '已取消收藏', 'is_favorite': False})
    else:
        cursor.execute("INSERT INTO user_company_favorites (user_id, company_id) VALUES (?,?)", (current_user_id, company_id))
        conn.commit()
        conn.close()
        return jsonify({'message': '已添加收藏', 'is_favorite': True})

@app.route('/api/companies/<int:company_id>', methods=['DELETE', 'OPTIONS'])
@token_required
def delete_company(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        logger.info('Company delete denied: user_id=%s company_id=%s', current_user_id, company_id)
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("SELECT name FROM companies WHERE id=?", (company_id,))
    row = cursor.fetchone()
    company_name = row[0] if row else str(company_id)
    # 与公司严格绑定：删除该公司相关的警报、媒体评价、相关资讯、关键词、用户关联与收藏，再删公司
    cursor.execute("DELETE FROM risk_alerts WHERE company_id=?", (company_id,))
    cursor.execute("DELETE FROM company_media_reviews WHERE company_id=?", (company_id,))
    cursor.execute("DELETE FROM company_news WHERE company_id=?", (company_id,))
    cursor.execute("DELETE FROM company_keywords WHERE company_id=?", (company_id,))
    cursor.execute("DELETE FROM user_company_favorites WHERE company_id=?", (company_id,))
    cursor.execute("DELETE FROM user_companies WHERE company_id=?", (company_id,))
    cursor.execute("DELETE FROM companies WHERE id=?", (company_id,))
    conn.commit()
    _audit_log(current_user_id, 'delete_company', 'company', company_id, company_name, cursor_=cursor, conn=conn)
    conn.close()
    return jsonify({'message': '删除成功'})

@app.route('/api/companies/<int:company_id>/crawl', methods=['POST', 'OPTIONS'])
@token_required
def trigger_company_crawl(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        logger.info('Company crawl denied: user_id=%s company_id=%s', current_user_id, company_id)
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("SELECT name FROM companies WHERE id=?", (company_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'message': 'Company not found'}), 404
    # 与添加企业一致：后台执行工商信息 + 相关新闻搜索 + 媒体舆情爬取
    t = threading.Thread(target=_run_company_crawl_and_news_pipeline, args=(company_id, row[0], current_user_id))
    t.daemon = True
    t.start()
    return jsonify({'message': '爬取任务已启动（含工商信息、相关新闻与媒体舆情）', 'crawl_status': 'started'})

@app.route('/api/companies/<int:company_id>/news-search', methods=['POST', 'OPTIONS'])
@token_required
def search_company_news_api(current_user_id, company_id):
    """大模型联网搜索企业相关新闻，整理后保存到 company_news"""
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM companies WHERE id=?", (company_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cursor.execute("SELECT api_key, base_url, model, enable_web_search FROM llm_config WHERE user_id=?", (current_user_id,))
    llm_row = cursor.fetchone()
    conn.close()
    if not llm_row or not (llm_row[0] and str(llm_row[0]).strip()):
        return jsonify({'message': '请先在系统设置中配置 LLM API 并开启联网搜索'}), 400
    api_key, base_url, model = llm_row[0], llm_row[1] or '', llm_row[2] or 'gpt-4o-mini'
    enable_web_search = bool(llm_row[3]) if len(llm_row) > 3 else False
    db_path = _DB_PATH
    conn0 = _db_conn()
    cur0 = conn0.cursor()
    cur0.execute("INSERT INTO task_runs (task_type, company_id, status, message) VALUES (?,?,?,?)", ('news_search', company_id, 'running', '大模型搜索相关新闻'))
    conn0.commit()
    task_run_id = cur0.lastrowid
    conn0.close()
    try:
        from backend.services.llm_service import search_company_news
        news_list = search_company_news(row[0], api_key, base_url, model, enable_web_search=enable_web_search)
        conn2 = sqlite3.connect(db_path)
        cur = conn2.cursor()
        cur.execute("UPDATE task_runs SET status=?, message=?, finished_at=? WHERE id=?", ('success', f'共{len(news_list)}条', datetime.utcnow().isoformat(), task_run_id))
        # 严格替换：先清空该企业原有相关新闻，再写入本次搜索结果，避免在原基础上追加导致混入别家/旧数据
        cur.execute("DELETE FROM company_news WHERE company_id=?", (company_id,))
        company_name_tag = (row[0] or '').strip()
        for n in news_list:
            rd = n.get('risk_dimensions') or {}
            rl = n.get('risk_level', '中')
            cur.execute("""INSERT INTO company_news (company_id, company_tag, title, content, source, sentiment_score, risk_level, category, publish_date, risk_dimensions)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (company_id, company_name_tag, n.get('title', ''), n.get('summary', ''), n.get('source', ''),
                 0.5, rl, n.get('category', '其他'), n.get('date', ''),
                 json.dumps(rd, ensure_ascii=False)))
            if (rl or '').strip() == '高':
                _create_risk_alert(cur, company_id, 'news', 'high', (n.get('title') or '')[:200], n.get('source') or '')
                _notify_alert_created(conn2, company_id, 'news', 'high', (n.get('title') or '')[:200])
        try:
            cur.execute("UPDATE companies SET news_status='success' WHERE id=?", (company_id,))
        except sqlite3.OperationalError:
            pass
        conn2.commit()
        conn2.close()
        _sync_risk_pipeline_after_news(company_id)
        return jsonify({'message': f'已更新该企业相关新闻，共 {len(news_list)} 条', 'count': len(news_list), 'news': news_list})
    except Exception as e:
        import traceback
        conn_err = _db_conn()
        conn_err.cursor().execute("UPDATE task_runs SET status=?, message=?, finished_at=? WHERE id=?", ('error', str(e)[:200], datetime.utcnow().isoformat(), task_run_id))
        conn_err.commit()
        conn_err.close()
        print('News search error:', traceback.format_exc())
        return jsonify({'message': str(e)}), 500

@app.route('/api/companies/<int:company_id>/news', methods=['DELETE', 'OPTIONS'])
@token_required
def clear_company_news(current_user_id, company_id):
    """清空该企业全部相关新闻（按企业专属标签：只删本企业标签的新闻，或按 company_id 全删）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("SELECT name FROM companies WHERE id=?", (company_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': '企业不存在'}), 404
    company_name = (row[0] or '').strip()
    cursor.execute("DELETE FROM company_news WHERE company_id=? AND (company_tag IS NULL OR company_tag=?)", (company_id, company_name))
    deleted = cursor.rowcount
    conn.commit()
    _audit_log(current_user_id, 'clear_company_news', 'company', company_id, f'清空{deleted}条', cursor_=cursor, conn=conn)
    conn.close()
    return jsonify({'message': f'已清空该企业相关新闻，共删除 {deleted} 条', 'deleted': deleted})

@app.route('/api/companies/<int:company_id>/export', methods=['GET', 'OPTIONS'])
@token_required
def export_company_report(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("""SELECT name, industry, risk_level, legal_representative, registered_capital, business_status,
        registered_address, business_scope, equity_structure, social_evaluation, established_date, last_updated
        FROM companies WHERE id=?""", (company_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cursor.execute("SELECT platform, title, content, sentiment FROM company_media_reviews WHERE company_id=?", (company_id,))
    reviews = cursor.fetchall()
    conn.close()
    # 固定格式文档（文本形式，便于扩展为 PDF/DOCX）
    lines = [
        '=' * 60,
        f'企业风险评估报告 - {row[0]}',
        '=' * 60,
        '',
        '【基本信息】',
        f'企业名称：{row[0]}',
        f'所属行业：{row[1] or "-"}',
        f'风险等级：{row[2] or "-"}',
        f'法定代表人：{row[3] or "-"}',
        f'注册资本：{row[4] or "-"}',
        f'经营状态：{row[5] or "-"}',
        f'注册地址：{row[6] or "-"}',
        f'经营范围：{row[7] or "-"}',
        f'股权结构：{row[8] or "-"}',
        f'社会评价：{row[9] or "-"}',
        f'成立日期：{row[10] or "-"}',
        f'更新时间：{row[11] or "-"}',
        '',
        '【媒体舆情】',
    ]
    for r in reviews:
        lines.append(f'  [{r[0]}] {r[1]} - {r[3]}')
        lines.append(f'    {r[2][:200]}...' if len(r[2] or '') > 200 else f'    {r[2] or ""}')
    lines.extend(['', '=' * 60, f'报告生成时间：{datetime.now().isoformat()}', '=' * 60])
    text = '\n'.join(lines)
    _audit_log(current_user_id, 'export_report', 'company', company_id, 'TXT')
    from flask import Response
    return Response(text, mimetype='text/plain; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=report_{company_id}.txt'})


@app.route('/api/companies/<int:company_id>/export-excel', methods=['GET', 'OPTIONS'])
@token_required
def export_company_excel(current_user_id, company_id):
    """导出企业报告为 Excel（多 Sheet：基本信息、媒体舆情、相关新闻、尽调补充）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'message': '服务端未安装 openpyxl，请使用 TXT 导出'}), 503
    from flask import send_file
    import io
    db_path = _DB_PATH
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("""SELECT name, industry, risk_level, legal_representative, registered_capital, business_status,
        registered_address, business_scope, equity_structure, social_evaluation, established_date, last_updated, supplement_notes
        FROM companies WHERE id=?""", (company_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    try:
        supplement_notes = row[12] if len(row) > 12 else ''
    except IndexError:
        supplement_notes = ''
    cursor.execute("SELECT platform, title, content, sentiment FROM company_media_reviews WHERE company_id=?", (company_id,))
    reviews = cursor.fetchall()
    cursor.execute("""SELECT title, content, source, risk_level, category, publish_date, risk_dimensions
        FROM company_news WHERE company_id=? ORDER BY created_at DESC LIMIT 200""", (company_id,))
    news_rows = cursor.fetchall()
    cursor.execute("SELECT role, content, merged_fields, created_at FROM company_supplements WHERE company_id=? ORDER BY created_at ASC", (company_id,))
    supplements = cursor.fetchall()
    conn.close()

    wb = Workbook()
    basic = wb.active
    basic.title = '基本信息'
    basic.append(['字段', '内容'])
    basic.append(['企业名称', row[0] or ''])
    basic.append(['行业', row[1] or ''])
    basic.append(['风险等级', row[2] or ''])
    basic.append(['法定代表人', row[3] or ''])
    basic.append(['注册资本', row[4] or ''])
    basic.append(['经营状态', row[5] or ''])
    basic.append(['注册地址', row[6] or ''])
    basic.append(['经营范围', row[7] or ''])
    basic.append(['股权结构', row[8] or ''])
    basic.append(['社会评价', row[9] or ''])
    basic.append(['成立日期', row[10] or ''])
    basic.append(['更新时间', row[11] or ''])
    basic.append(['尽调备注', supplement_notes or ''])

    rev_ws = wb.create_sheet('媒体舆情')
    rev_ws.append(['平台', '标题', '内容摘要', '情感'])
    for r in reviews:
        rev_ws.append([r[0] or '', (r[1] or '')[:200], (r[2] or '')[:500], r[3] or ''])

    news_ws = wb.create_sheet('相关新闻')
    news_ws.append(['标题', '内容摘要', '来源', '风险等级', '分类', '发布日期', '风险维度'])
    for r in news_rows:
        rd = r[6] if len(r) > 6 and r[6] else ''
        if isinstance(rd, str) and rd:
            try:
                rd = json.loads(rd)
                rd = json.dumps(rd, ensure_ascii=False) if rd else ''
            except Exception:
                pass
        news_ws.append([r[0] or '', (r[1] or '')[:500], r[2] or '', r[3] or '', r[4] or '', r[5] or '', rd])

    sup_ws = wb.create_sheet('尽调补充')
    sup_ws.append(['角色', '内容', '归集字段', '时间'])
    for s in supplements:
        sup_ws.append([s[0] or '', (s[1] or '')[:1000], s[2] or '', s[3] or ''])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _audit_log(current_user_id, 'export_report', 'company', company_id, 'Excel')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'report_{company_id}.xlsx')


@app.route('/api/companies/<int:company_id>/export-pdf', methods=['GET', 'OPTIONS'])
@token_required
def export_company_pdf(current_user_id, company_id):
    """导出企业报告为 PDF"""
    if request.method == 'OPTIONS':
        return jsonify({})
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return jsonify({'message': '服务端未安装 reportlab，请使用 TXT 或 Excel 导出'}), 503
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("""SELECT name, industry, risk_level, legal_representative, registered_capital, business_status,
        registered_address, business_scope, equity_structure, social_evaluation, established_date, last_updated, supplement_notes
        FROM companies WHERE id=?""", (company_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cursor.execute("SELECT platform, title, content, sentiment FROM company_media_reviews WHERE company_id=?", (company_id,))
    reviews = cursor.fetchall()
    cursor.execute("SELECT title, content, source, risk_level, category FROM company_news WHERE company_id=? ORDER BY created_at DESC LIMIT 50", (company_id,))
    news_rows = cursor.fetchall()
    conn.close()
    supplement_notes = row[12] if len(row) > 12 else ''
    from flask import send_file
    import io
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, f"Enterprise Risk Report - {row[0] or ''}")
    y -= 30
    c.setFont("Helvetica", 10)
    blocks = [
        ("Basic Info", [
            f"Name: {row[0] or '-'}", f"Industry: {row[1] or '-'}", f"Risk Level: {row[2] or '-'}",
            f"Legal Rep: {row[3] or '-'}", f"Capital: {row[4] or '-'}", f"Status: {row[5] or '-'}",
            f"Address: {row[6] or '-'}", f"Scope: {(row[7] or '-')[:200]}", f"Social: {row[9] or '-'}",
            f"Updated: {row[11] or '-'}", f"Supplement: {(supplement_notes or '-')[:300]}"
        ]),
        ("Media", [f"[{r[0]}] {r[1]} - {r[3]}" for r in reviews[:20]]),
        ("News", [f"{r[0]} | {r[2]} | {r[4]}" for r in news_rows[:15]])
    ]
    for title, lines in blocks:
        y -= 20
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, title)
        y -= 18
        c.setFont("Helvetica", 9)
        for line in lines:
            if y < 50:
                c.showPage()
                y = height - 40
            c.drawString(55, y, (line or '')[:90])
            y -= 14
    c.drawString(50, 30, f"Generated: {datetime.now().isoformat()}")
    c.save()
    buf.seek(0)
    _audit_log(current_user_id, 'export_report', 'company', company_id, 'PDF')
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=f'report_{company_id}.pdf')


@app.route('/api/llm-config', methods=['GET', 'OPTIONS'])
@token_required
def get_llm_config(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT api_key, base_url, model, enable_web_search FROM llm_config WHERE user_id=?", (current_user_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return jsonify({
            'api_key': row[0] or '', 'base_url': row[1] or '', 'model': row[2] or 'gpt-4o-mini',
            'enable_web_search': bool(row[3]) if len(row) > 3 else False
        })
    return jsonify({'api_key': '', 'base_url': '', 'model': 'gpt-4o-mini', 'enable_web_search': False})

@app.route('/api/llm-config', methods=['PUT', 'OPTIONS'])
@token_required
def put_llm_config(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    enable_web = 1 if data.get('enable_web_search') else 0
    cursor.execute("""INSERT OR REPLACE INTO llm_config (user_id, api_key, base_url, model, enable_web_search, updated_at)
        VALUES (?,?,?,?,?,?)""",
        (current_user_id, data.get('api_key',''), data.get('base_url',''), data.get('model','gpt-4o-mini'), enable_web, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'message': 'LLM 配置已保存'})

@app.route('/api/test-llm-company', methods=['POST', 'OPTIONS'])
@token_required
def test_llm_company(current_user_id):
    """调试用：测试 LLM 工商信息搜集是否正常，返回原始结果便于排查"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    name = data.get('name', '美团')
    conn = sqlite3.connect(_DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT api_key, base_url, model, enable_web_search FROM llm_config WHERE user_id=?", (current_user_id,))
    row = cur.fetchone()
    conn.close()
    if not row or not (row[0] and row[0].strip()):
        return jsonify({
            'success': False,
            'error': '未配置 LLM API',
            'hint': '请先在【系统设置】中填写 API Key、Base URL、模型，并勾选联网搜索'
        }), 400
    api_key, base_url, model = row[0], row[1] or '', row[2] or 'gpt-4o-mini'
    enable_web_search = bool(row[3]) if len(row) > 3 else False
    try:
        from backend.services.llm_service import fetch_company_info_by_llm
        result = fetch_company_info_by_llm(name, api_key, base_url, model, enable_web_search=enable_web_search)
        if result:
            return jsonify({
                'success': True,
                'company_name': name,
                'result': result,
                'config_used': {'base_url': base_url[:50] + '...' if len(base_url or '') > 50 else base_url, 'model': model, 'enable_web_search': enable_web_search}
            })
        return jsonify({
            'success': False,
            'error': 'LLM 未返回有效 JSON',
            'hint': '请检查模型是否支持联网搜索（如通义千问需 qwen-plus 或 qwen3-max），或查看后端控制台 LLM call error'
        }), 400
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/system-status', methods=['GET', 'OPTIONS'])
@token_required
def get_system_status(current_user_id):
    """检测后端、API、爬虫等真实状态"""
    if request.method == 'OPTIONS':
        return jsonify({})
    services = []
    base_url = request.host_url.rstrip('/')
    # 1. 后端 API（能执行到此说明后端在线）
    services.append({'id': 'backend', 'name': '后端 API', 'endpoint': base_url, 'status': 'online'})
    # 2. 数据库
    try:
        conn = sqlite3.connect(_DB_PATH)
        conn.execute("SELECT 1")
        conn.close()
        services.append({'id': 'database', 'name': '数据库', 'endpoint': 'SQLite', 'status': 'online'})
    except Exception as e:
        services.append({'id': 'database', 'name': '数据库', 'endpoint': 'SQLite', 'status': 'offline', 'message': str(e)[:50]})
    # 3. 企业信息（LLM/爬虫）
    try:
        conn = sqlite3.connect(_DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT api_key, enable_web_search FROM llm_config WHERE api_key IS NOT NULL AND api_key != '' LIMIT 1")
        row = cur.fetchone()
        conn.close()
        if row:
            en_web = bool(row[1]) if len(row) > 1 else False
            services.append({'id': 'llm', 'name': '大模型（工商信息）', 'endpoint': 'LLM API', 'status': 'online' if en_web else 'warning', 'message': '已配置并开启联网' if en_web else '已配置但未开启联网搜索'})
        else:
            services.append({'id': 'llm', 'name': '大模型（工商信息）', 'endpoint': '未配置', 'status': 'warning', 'message': '请在系统设置中配置 API'})
    except Exception as e:
        services.append({'id': 'llm', 'name': '大模型', 'endpoint': '-', 'status': 'offline', 'message': str(e)[:50]})
    # 4. MediaCrawler
    try:
        from backend.services.mediacrawler_service import is_available
        av = is_available()
        services.append({'id': 'mediacrawler', 'name': '媒体爬虫', 'endpoint': 'MediaCrawler', 'status': 'online' if av else 'offline', 'message': '已配置' if av else '请设置 MEDIACRAWLER_PATH'})
    except Exception as e:
        services.append({'id': 'mediacrawler', 'name': '媒体爬虫', 'endpoint': 'MediaCrawler', 'status': 'offline', 'message': str(e)[:50]})
    # 5. 小红书登录态（与平台登录一致：logged_in.txt 或 session 文件）
    try:
        sessions_dir = _mediacrawler_sessions_dir()
        path = os.path.join(sessions_dir, 'xhs')
        marker_file = os.path.join(path, 'logged_in.txt')
        ignore_files = {'login_stderr.log'}
        dir_has = False
        if os.path.isdir(path):
            dir_has = any(f for f in os.listdir(path) if not f.startswith('.') and f not in ignore_files)
        has_xhs = os.path.isfile(marker_file) or dir_has
        services.append({'id': 'xhs-login', 'name': '小红书登录', 'endpoint': '扫码/上传 Cookie', 'status': 'online' if has_xhs else 'offline', 'message': '已登录，可爬取小红书' if has_xhs else '请到系统设置→平台登录中扫码或上传 Cookie'})
    except Exception as e:
        services.append({'id': 'xhs-login', 'name': '小红书登录', 'endpoint': '-', 'status': 'offline', 'message': str(e)[:50]})
    return jsonify({'services': services})

@app.route('/api/mediacrawler/status', methods=['GET', 'OPTIONS'])
@token_required
def get_mediacrawler_status(current_user_id):
    """检查 MediaCrawler 是否已配置并可用"""
    if request.method == 'OPTIONS':
        return jsonify({})
    try:
        from backend.services.mediacrawler_service import is_available, get_mediacrawler_path, get_mediacrawler_platform
        available = is_available()
        return jsonify({
            'available': available,
            'path': get_mediacrawler_path() if available else '',
            'platform': get_mediacrawler_platform(),
            'message': 'MediaCrawler 已配置' if available else '请设置环境变量 MEDIACRAWLER_PATH 指向 MediaCrawler 项目根目录'
        })
    except Exception as e:
        return jsonify({'available': False, 'path': '', 'platform': 'xhs', 'message': str(e)})


@app.route('/api/mediacrawler/config', methods=['GET', 'POST', 'OPTIONS'])
@token_required
def mediacrawler_config(current_user_id):
    """GET 返回当前 MediaCrawler 路径；POST 保存路径（.app 下写入应用数据目录 .env）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    if request.method == 'GET':
        path = os.environ.get('MEDIACRAWLER_PATH', '').strip()
        return jsonify({'path': path})
    # POST
    data = request.get_json() or {}
    path = (data.get('path') or '').strip()
    path_ok = path and os.path.isdir(path) and os.path.isfile(os.path.join(path, 'main.py'))
    if _standalone_data:
        _data_env = os.path.join(_DATA_DIR, '.env')
        try:
            _lines = []
            if os.path.isfile(_data_env):
                with open(_data_env, 'r', encoding='utf-8') as f:
                    for line in f:
                        if not line.strip().startswith('MEDIACRAWLER_PATH='):
                            _lines.append(line.rstrip('\n'))
            _lines.append('MEDIACRAWLER_PATH=%s' % path.replace('\\', '/'))
            with open(_data_env, 'w', encoding='utf-8') as f:
                f.write('\n'.join(_lines) + '\n')
        except Exception as e:
            return jsonify({'ok': False, 'message': '写入配置失败: %s' % e}), 500
    os.environ['MEDIACRAWLER_PATH'] = path
    return jsonify({'ok': True, 'path': path, 'valid': path_ok, 'message': '已保存；路径有效，可使用媒体爬虫' if path_ok else '已保存；请确认路径指向 MediaCrawler 项目根目录（含 main.py）'})


def _mediacrawler_sessions_dir():
    from backend.services.mediacrawler_service import get_sessions_dir
    return get_sessions_dir()


def _mediacrawler_qr_dir():
    from backend.services.mediacrawler_service import get_qr_dir
    return get_qr_dir()


@app.route('/api/mediacrawler/clear-login', methods=['POST', 'OPTIONS'])
@token_required
def mediacrawler_clear_login(current_user_id):
    """清除当前平台扫码登录态，便于重新扫码（如换号）。会删除 marker 并重命名 browser_data 目录。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    platform = (data.get('platform') or 'xhs').strip() or 'xhs'
    sessions_dir = _mediacrawler_sessions_dir()
    platform_session = os.path.join(sessions_dir, platform)
    marker = os.path.join(platform_session, 'logged_in.txt')
    try:
        if os.path.isfile(marker):
            os.remove(marker)
    except Exception:
        pass
    mc_path = os.environ.get('MEDIACRAWLER_PATH', '').strip()
    if mc_path and os.path.isdir(mc_path):
        browser_data_dir = os.path.join(mc_path, 'browser_data', f'{platform}_user_data_dir')
        if os.path.isdir(browser_data_dir):
            try:
                bak = browser_data_dir + '.bak.' + datetime.now().strftime('%Y%m%d%H%M%S')
                os.rename(browser_data_dir, bak)
            except Exception:
                pass
    return jsonify({'ok': True, 'message': '已清除当前平台登录态，可重新点击「生成二维码」扫码', 'platform': platform})


@app.route('/api/mediacrawler/start-login', methods=['POST', 'OPTIONS'])
@token_required
def mediacrawler_start_login(current_user_id):
    """后台启动 MediaCrawler 扫码登录子进程，二维码将写入 data/mediacrawler_qr/{platform}.txt，前端轮询 login-qr 与 login-status。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    platform = (data.get('platform') or 'xhs').strip() or 'xhs'
    force_new = data.get('force_new') is True  # 为 true 时先清除登录态再启动，确保会出二维码
    mc_path = os.environ.get('MEDIACRAWLER_PATH', '').strip()
    if not mc_path or not os.path.isdir(mc_path):
        return jsonify({'started': False, 'message': 'MediaCrawler 未配置（请设置 MEDIACRAWLER_PATH）'}), 400
    qr_dir = _mediacrawler_qr_dir()
    qr_file = os.path.abspath(os.path.join(qr_dir, f'{platform}.txt'))
    sessions_dir = os.path.abspath(_mediacrawler_sessions_dir())
    script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'scripts', 'run_media_login_with_qr.py'))
    # frozen 时用 -m backend.scripts.run_media_login_with_qr，脚本在 PYZ 内，无需磁盘上的 .py 文件
    if not getattr(sys, 'frozen', False) and not os.path.isfile(script_path):
        return jsonify({'started': False, 'message': '登录脚本不存在'}), 500
    if force_new:
        platform_session = os.path.join(sessions_dir, platform)
        marker = os.path.join(platform_session, 'logged_in.txt')
        try:
            if os.path.isfile(marker):
                os.remove(marker)
        except Exception:
            pass
        browser_data_dir = os.path.join(mc_path, 'browser_data', f'{platform}_user_data_dir')
        if os.path.isdir(browser_data_dir):
            try:
                bak = browser_data_dir + '.bak.' + datetime.now().strftime('%Y%m%d%H%M%S')
                os.rename(browser_data_dir, bak)
            except Exception:
                pass
    # 清空旧二维码，便于轮询判断「尚未生成」
    try:
        if os.path.isfile(qr_file):
            os.remove(qr_file)
    except Exception:
        pass
    env = dict(os.environ)
    env['MEDIACRAWLER_PATH'] = os.path.abspath(mc_path)
    env['MEDIACRAWLER_QR_FILE'] = qr_file
    env['SYS2_SESSIONS_DIR'] = sessions_dir
    env['PLATFORM'] = platform
    # .app 内用同一可执行文件跑登录脚本；Playwright 浏览器装到应用数据目录，首次自动 install
    if getattr(sys, 'frozen', False):
        _pw_browsers = os.path.join(_DATA_DIR, 'playwright-browsers')
        env['PLAYWRIGHT_BROWSERS_PATH'] = _pw_browsers
        # macOS 上新 headless 易崩溃（mach_vm_read / crash info version 7），强制用旧模式
        env['PLAYWRIGHT_CHROMIUM_USE_HEADLESS_NEW'] = '0'
        _pw_marker = os.path.join(_DATA_DIR, '.playwright_installed')
        _chromium_ok = os.path.isdir(_pw_browsers) and any(
            'chromium' in n.lower() for n in (os.listdir(_pw_browsers) if os.path.isdir(_pw_browsers) else [])
        )
        if not _chromium_ok or not os.path.isfile(_pw_marker):
            try:
                proc = subprocess.run(
                    [sys.executable, '-m', 'playwright', 'install', 'chromium'],
                    env=env, timeout=180, capture_output=True, text=True
                )
                _chromium_ok = proc.returncode == 0 and os.path.isdir(_pw_browsers)
                if _chromium_ok:
                    with open(_pw_marker, 'w') as _f:
                        _f.write('1')
                elif proc.returncode != 0 and proc.stderr:
                    logger.warning('playwright install chromium failed: %s', proc.stderr[:500])
            except Exception as e:
                logger.warning('playwright install chromium exception: %s', e)
            _chromium_ok = _chromium_ok or (os.path.isdir(_pw_browsers) and any(
                'chromium' in n.lower() for n in (os.listdir(_pw_browsers) if os.path.isdir(_pw_browsers) else [])
            ))
        if not _chromium_ok:
            return jsonify({
                'started': False,
                'message': 'Chromium 未安装成功，无法生成二维码。请确保网络正常后重试；或在本机终端执行：PLAYWRIGHT_BROWSERS_PATH="%s" "%s" -m playwright install chromium' % (_pw_browsers, sys.executable)
            }), 400
        env['RISKGUARD_LOGIN_SUBPROCESS'] = '1'
        run_argv = [sys.executable, '-m', 'backend.scripts.run_media_login_with_qr']
    else:
        from backend.services.mediacrawler_service import get_mediacrawler_python
        py_exe = get_mediacrawler_python()
        if not py_exe or not os.path.isfile(py_exe):
            py_exe = 'python3'
        run_argv = [py_exe, script_path]
    platform_session = os.path.join(sessions_dir, platform)
    os.makedirs(platform_session, exist_ok=True)
    stderr_log = os.path.join(platform_session, "login_stderr.log")
    try:
        with open(stderr_log, "w", encoding="utf-8") as stderr_f:
            proc = subprocess.Popen(
                run_argv,
                cwd=os.path.abspath(mc_path),
                env=env,
                stdout=subprocess.DEVNULL,
                stderr=stderr_f,
            )
    except Exception as e:
        return jsonify({'started': False, 'message': str(e)}), 500
    return jsonify({'started': True, 'message': '已启动登录流程，请轮询二维码与登录状态', 'platform': platform})


@app.route('/api/mediacrawler/login-qr', methods=['GET', 'OPTIONS'])
@token_required
def mediacrawler_login_qr(current_user_id):
    """读取当前平台的二维码 Base64 内容（由 start-login 子进程写入），供前端轮询显示。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    platform = request.args.get('platform', 'xhs').strip() or 'xhs'
    qr_dir = _mediacrawler_qr_dir()
    qr_file = os.path.abspath(os.path.join(qr_dir, f'{platform}.txt'))
    qrcode = None
    if os.path.isfile(qr_file):
        try:
            with open(qr_file, 'r', encoding='utf-8') as f:
                content = f.read().strip()
            if content:
                if content.startswith('data:'):
                    qrcode = content
                else:
                    qrcode = 'data:image/png;base64,' + content
        except Exception:
            pass
    return jsonify({'platform': platform, 'qrcode': qrcode})


@app.route('/api/mediacrawler/login-status', methods=['GET', 'OPTIONS'])
@token_required
def mediacrawler_login_status(current_user_id):
    """查询指定平台是否已有登录态（已上传 session/cookie）。?debug=1 时返回检测路径便于排查。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    platform = request.args.get('platform', 'xhs').strip() or 'xhs'
    sessions_dir = _mediacrawler_sessions_dir()
    path = os.path.join(sessions_dir, platform)
    marker_file = os.path.join(path, 'logged_in.txt')
    # 仅以 logged_in.txt（扫码成功）或其它 session/cookie 文件为登录态，排除本系统写的调试日志
    ignore_files = {'login_stderr.log'}
    dir_has_session_file = False
    if os.path.isdir(path):
        dir_has_session_file = any(f for f in os.listdir(path) if not f.startswith('.') and f not in ignore_files)
    has_session = os.path.isfile(marker_file) or dir_has_session_file or os.path.isfile(path + '.json') or os.path.isfile(path)
    out = {
        'platform': platform,
        'status': 'has_session' if has_session else 'no_session',
        'message': '已存在登录态，可直接爬取' if has_session else '请在本机扫码后上传 Cookie/Session，或使用下方上传'
    }
    if request.args.get('debug'):
        debug_info = {
            'sessions_dir': sessions_dir,
            'marker_file': marker_file,
            'marker_exists': os.path.isfile(marker_file),
            'path_exists': os.path.isdir(path),
        }
        stderr_log = os.path.join(path, 'login_stderr.log')
        if os.path.isfile(stderr_log):
            try:
                with open(stderr_log, 'r', encoding='utf-8', errors='replace') as f:
                    tail = f.read()[-2000:]
                debug_info['login_stderr_tail'] = tail
            except Exception:
                pass
        out['_debug'] = debug_info
    return jsonify(out)


@app.route('/api/mediacrawler/upload-session', methods=['POST', 'OPTIONS'])
@token_required
def mediacrawler_upload_session(current_user_id):
    """上传 MediaCrawler 某平台的 Cookie/Session 文件（本机扫码后从 MediaCrawler 数据目录复制）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    platform = (request.form.get('platform') or 'xhs').strip() or 'xhs'
    if 'file' not in request.files:
        return jsonify({'message': '请选择要上传的文件'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'message': '请选择文件'}), 400
    allowed_ext = ('.json', '.txt')
    base = os.path.splitext(f.filename)[0]
    ext = (os.path.splitext(f.filename)[1] or '').lower()
    if ext not in allowed_ext:
        return jsonify({'message': '仅支持 .json 或 .txt 文件'}), 400
    f.seek(0, 2)
    size = f.tell()
    f.seek(0)
    if size > 2 * 1024 * 1024:
        return jsonify({'message': '文件大小不超过 2MB'}), 400
    sessions_dir = _mediacrawler_sessions_dir()
    platform_dir = os.path.join(sessions_dir, platform)
    os.makedirs(platform_dir, exist_ok=True)
    safe_name = (base or 'cookies')[:80] + ext
    dest = os.path.join(platform_dir, safe_name)
    try:
        f.save(dest)
        marker = os.path.join(platform_dir, 'logged_in.txt')
        try:
            with open(marker, 'w', encoding='utf-8') as mf:
                mf.write('ok\n')
        except Exception:
            pass
        return jsonify({'message': f'已保存为 {platform}/{safe_name}，已标记登录态，爬虫将使用该登录态', 'platform': platform})
    except Exception as e:
        return jsonify({'message': str(e)}), 500


@app.route('/api/mediacrawler/login-qrcode', methods=['GET', 'OPTIONS'])
@token_required
def mediacrawler_login_qrcode(current_user_id):
    """云服务器无法直接弹窗扫码，建议：本机执行一次扫码后上传 Cookie；或后续接入可输出二维码的登录流程。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    return jsonify({
        'supported': False,
        'message': '当前服务端无法生成二维码。请在本机运行 MediaCrawler 扫码登录后，将数据目录中的 Cookie/Session 文件通过「上传 Cookie」上传到此处。'
    }), 200


@app.route('/api/mediacrawler/test-crawl', methods=['POST', 'OPTIONS'])
@token_required
def mediacrawler_test_crawl(current_user_id):
    """测试媒体爬虫：用关键词（默认美团）执行一次简短爬取并返回少量数据。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    keyword = (data.get('keyword') or '美团').strip() or '美团'
    platform = (data.get('platform') or 'xhs').strip() or 'xhs'
    try:
        from backend.services.mediacrawler_service import is_available, crawl_by_keyword
        if not is_available():
            return jsonify({
                'ok': False,
                'message': 'MediaCrawler 未配置（请设置 MEDIACRAWLER_PATH）',
                'sample': [
                    {'platform': platform, 'title': '（模拟）测试条目', 'content': '配置 MEDIACRAWLER_PATH 后可进行真实爬取测试。'}
                ]
            })
        result = crawl_by_keyword(keyword, platform=platform, callback=None, company_id=None)
        status = result.get('status', '')
        reviews = result.get('reviews', [])[:5]
        # 未登录或未解析到内容时 status 为 no_content，不显示为「爬取成功」
        ok = status in ('ok', 'mock')
        out = {
            'ok': ok,
            'keyword': keyword,
            'platform': platform,
            'message': result.get('message', ''),
            'count': len(result.get('reviews', [])) if ok else 0,
            'sample': [{'platform': r.get('platform', platform), 'title': (r.get('title') or '')[:80], 'content': (r.get('content') or '')[:200]} for r in reviews]
        }
        if result.get('crawl_stderr'):
            out['crawl_stderr'] = result.get('crawl_stderr')
        return jsonify(out)
    except Exception as e:
        return jsonify({
            'ok': False,
            'message': str(e)[:200],
            'sample': []
        })


@app.route('/api/news/rolling', methods=['GET', 'OPTIONS'])
@token_required
def get_rolling_news(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    limit = int(request.args.get('limit', 50))
    company_id_raw = request.args.get('company_id')
    # 严格按企业过滤：仅当 company_id 为有效整数时才过滤，否则若带了参数但无效则返回空，避免误返回全部
    filter_company_id = None
    if company_id_raw is not None and company_id_raw != '':
        try:
            filter_company_id = int(company_id_raw)
        except (TypeError, ValueError):
            return jsonify([])
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    try:
        cids = _user_company_ids(cursor, current_user_id)
        if filter_company_id is not None:
            if filter_company_id not in cids:
                conn.close()
                return jsonify([])
            # 按企业专属标签筛选：只返回 company_id 匹配且 company_tag 与当前企业名一致的新闻（无标签的旧数据不再展示）
            cursor.execute("""SELECT id, company_id, title, content, source, source_url, sentiment_score, risk_level, category, publish_date, created_at, risk_dimensions
                FROM company_news WHERE company_id=? AND company_tag = (SELECT name FROM companies WHERE id=?)
                ORDER BY created_at DESC LIMIT ?""", (filter_company_id, filter_company_id, limit))
        else:
            if not cids:
                conn.close()
                return jsonify([])
            placeholders = ','.join('?' * len(cids))
            cursor.execute("""SELECT id, company_id, title, content, source, source_url, sentiment_score, risk_level, category, publish_date, created_at, risk_dimensions
                FROM company_news WHERE company_id IN (""" + placeholders + """) ORDER BY created_at DESC LIMIT ?""", tuple(cids) + (limit,))
        rows = cursor.fetchall()
        news = [{'id': r[0], 'company_id': r[1], 'company_name': '', 'title': r[2], 'content': r[3],
                 'source': r[4], 'source_url': r[5], 'sentiment_score': r[6], 'risk_level': r[7], 'category': r[8],
                 'publish_date': r[9], 'created_at': r[10], 'risk_dimensions': r[11] if len(r) > 11 and r[11] else None} for r in rows]
        # 按企业过滤时只保留 company_id 与请求一致的条目，防止表内错误数据导致串企业
        if filter_company_id is not None:
            news = [n for n in news if n.get('company_id') == filter_company_id]
    except sqlite3.OperationalError:
        if filter_company_id is not None:
            cursor.execute("""SELECT id, company_id, title, content, source, source_url, sentiment_score, risk_level, category, publish_date, created_at
                FROM company_news WHERE company_id=? AND company_tag = (SELECT name FROM companies WHERE id=?)
                ORDER BY created_at DESC LIMIT ?""", (filter_company_id, filter_company_id, limit))
            rows = cursor.fetchall()
        else:
            if cids:
                ph = ','.join('?' * len(cids))
                cursor.execute("""SELECT id, company_id, title, content, source, source_url, sentiment_score, risk_level, category, publish_date, created_at
                    FROM company_news WHERE company_id IN (""" + ph + """) ORDER BY created_at DESC LIMIT ?""", tuple(cids) + (limit,))
                rows = cursor.fetchall()
            else:
                rows = []
        news = [{'id': r[0], 'company_id': r[1], 'company_name': '', 'title': r[2], 'content': r[3],
                 'source': r[4], 'source_url': r[5], 'sentiment_score': r[6], 'risk_level': r[7], 'category': r[8],
                 'publish_date': r[9], 'created_at': r[10], 'risk_dimensions': None} for r in rows]
        if filter_company_id is not None:
            news = [n for n in news if n.get('company_id') == filter_company_id]
    cursor.execute("SELECT id, name FROM companies")
    company_map = {r[0]: r[1] for r in cursor.fetchall()}
    for n in news:
        n['company_name'] = company_map.get(n['company_id'], '')
    conn.close()
    return jsonify(news)

@app.route('/api/dashboard', methods=['GET', 'OPTIONS'])
@token_required
def get_dashboard(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify({'company_count': 0, 'news_today': 0, 'recent_companies': []})
    placeholders = ','.join('?' * len(cids))
    cursor.execute("SELECT COUNT(*) FROM companies WHERE id IN (" + placeholders + ")", tuple(cids))
    company_count = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM company_news WHERE company_id IN (" + placeholders + ") AND created_at > datetime('now','-1 day')", tuple(cids))
    news_today = cursor.fetchone()[0]
    cursor.execute("SELECT id, name, risk_level, last_updated FROM companies WHERE id IN (" + placeholders + ") ORDER BY last_updated DESC LIMIT 10", tuple(cids))
    recent = [{'id': r[0], 'name': r[1], 'risk_level': r[2], 'last_updated': r[3]} for r in cursor.fetchall()]
    conn.close()
    return jsonify({'company_count': company_count, 'news_today': news_today, 'recent_companies': recent})

@app.route('/api/dashboard/trend', methods=['GET', 'OPTIONS'])
@token_required
def get_dashboard_trend(current_user_id):
    """近 N 天每日资讯数量趋势（用于图表），仅统计当前用户企业下的资讯。支持 ?days=7|30|90"""
    if request.method == 'OPTIONS':
        return jsonify({})
    days = min(90, max(1, int(request.args.get('days', 7))))
    conn = _db_conn()
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        today = datetime.now().date()
        trend = [{'date': (today - timedelta(days=days-1-i)).isoformat(), 'count': 0, 'label': ['一','二','三','四','五','六','日'][(today - timedelta(days=days-1-i)).weekday()]} for i in range(days)]
        return jsonify({'trend': trend})
    placeholders = ','.join('?' * len(cids))
    cursor.execute("""SELECT date(created_at) as d, COUNT(*) FROM company_news 
        WHERE company_id IN (""" + placeholders + """) AND created_at > datetime('now','-""" + str(days) + """ day') GROUP BY date(created_at) ORDER BY d""", tuple(cids))
    rows = cursor.fetchall()
    conn.close()
    day_map = {r[0]: r[1] for r in rows}
    today = datetime.now().date()
    trend = []
    for i in range(days):
        d = (today - timedelta(days=days-1-i)).isoformat()
        trend.append({'date': d, 'count': day_map.get(d, 0), 'label': ['一','二','三','四','五','六','日'][(today - timedelta(days=days-1-i)).weekday()]})
    return jsonify({'trend': trend})


@app.route('/api/dashboard/risk-distribution', methods=['GET', 'OPTIONS'])
@token_required
def get_dashboard_risk_distribution(current_user_id):
    """企业风险等级分布（用于仪表盘图表），仅统计当前用户企业"""
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify({'低': 0, '中': 0, '高': 0, '未知': 0})
    placeholders = ','.join('?' * len(cids))
    cursor.execute("SELECT risk_level, COUNT(*) FROM companies WHERE id IN (" + placeholders + ") GROUP BY risk_level", tuple(cids))
    rows = cursor.fetchall()
    conn.close()
    dist = {'低': 0, '中': 0, '高': 0, '未知': 0}
    for r in rows:
        k = (r[0] or '未知').strip() or '未知'
        dist[k] = dist.get(k, 0) + r[1]
    return jsonify(dist)


@app.route('/api/dashboard/category-distribution', methods=['GET', 'OPTIONS'])
@token_required
def get_dashboard_category_distribution(current_user_id):
    """资讯分类分布（用于仪表盘图表），仅统计当前用户企业下的资讯。支持 ?days=7|30|90"""
    if request.method == 'OPTIONS':
        return jsonify({})
    days = min(90, max(1, int(request.args.get('days', 30))))
    conn = _db_conn()
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify([])
    placeholders = ','.join('?' * len(cids))
    cursor.execute("""SELECT category, COUNT(*) FROM company_news 
        WHERE company_id IN (""" + placeholders + """) AND category IS NOT NULL AND category != '' AND created_at > datetime('now','-""" + str(days) + """ day') 
        GROUP BY category ORDER BY COUNT(*) DESC""", tuple(cids))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([{'category': r[0], 'count': r[1]} for r in rows])


@app.route('/api/search', methods=['GET', 'OPTIONS'])
@token_required
def search(current_user_id):
    """搜索企业、资讯（仅限当前用户有权限的企业）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'companies': [], 'news': []})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify({'companies': [], 'news': []})
    placeholders = ','.join('?' * len(cids))
    cursor.execute("SELECT id, name, industry, risk_level FROM companies WHERE id IN (" + placeholders + ") AND (name LIKE ? OR industry LIKE ?) LIMIT 20",
        tuple(cids) + ('%' + q + '%', '%' + q + '%'))
    companies = [{'id': r[0], 'name': r[1], 'industry': r[2], 'risk_level': r[3]} for r in cursor.fetchall()]
    cursor.execute("""SELECT n.id, n.company_id, n.title, c.name FROM company_news n 
        LEFT JOIN companies c ON n.company_id = c.id 
        WHERE n.company_id IN (""" + placeholders + """) AND (n.title LIKE ? OR n.content LIKE ?) LIMIT 15""", tuple(cids) + ('%' + q + '%', '%' + q + '%'))
    news = [{'id': r[0], 'company_id': r[1], 'title': r[2], 'company_name': r[3]} for r in cursor.fetchall()]
    conn.close()
    return jsonify({'companies': companies, 'news': news})

@app.route('/api/companies/<int:company_id>/wordcloud', methods=['GET', 'OPTIONS'])
@token_required
def get_company_wordcloud(current_user_id, company_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT keyword, weight FROM company_keywords WHERE company_id=?", (company_id,))
    keywords = [{'word': r[0], 'weight': r[1]} for r in cursor.fetchall()]
    conn.close()
    try:
        from backend.services.wordcloud_service import generate_wordcloud
        img = generate_wordcloud(keywords)
        if img:
            return jsonify({'image': img})
    except Exception as e:
        print('Wordcloud error:', e)
    return jsonify({'image': None, 'keywords': keywords})


@app.route('/api/companies/<int:company_id>/supplements', methods=['GET', 'OPTIONS'])
@token_required
def get_company_supplements(current_user_id, company_id):
    """获取企业尽调补充对话历史"""
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限访问该企业'}), 403
    cursor.execute("SELECT id FROM companies WHERE id=?", (company_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    cursor.execute(
        "SELECT id, role, content, merged_fields, created_at FROM company_supplements WHERE company_id=? ORDER BY created_at ASC",
        (company_id,)
    )
    rows = cursor.fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'role': r[1], 'content': r[2], 'merged_fields': r[3], 'created_at': r[4]} for r in rows])


@app.route('/api/companies/<int:company_id>/supplement-chat', methods=['POST', 'OPTIONS'])
@token_required
def supplement_chat(current_user_id, company_id):
    """尽调补充：用户输入文字，LLM 整理后归集到企业并返回摘要"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'message': '请输入补充内容'}), 400
    if len(text) > 5000:
        return jsonify({'message': '补充内容不超过 5000 字'}), 400
    db_path = _DB_PATH
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute(
        "SELECT name, legal_representative, registered_capital, business_status, registered_address, business_scope, social_evaluation, supplement_notes FROM companies WHERE id=?",
        (company_id,)
    )
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    name = row[0]
    summary_parts = [f"企业名称：{name}"]
    for i, label in enumerate(['法定代表人', '注册资本', '经营状态', '注册地址', '经营范围', '社会评价', '内部备注'], 1):
        if i < len(row) and row[i]:
            summary_parts.append(f"{label}：{str(row[i])[:200]}")
    company_summary = '\n'.join(summary_parts)
    cursor.execute("SELECT api_key, base_url, model FROM llm_config WHERE user_id=?", (current_user_id,))
    llm_row = cursor.fetchone()
    conn.close()
    if not llm_row or not (llm_row[0] and str(llm_row[0]).strip()):
        return jsonify({'message': '请先在系统设置中配置 LLM API Key'}), 400
    from backend.services.llm_service import process_supplement_chat
    result = process_supplement_chat(name, company_summary, text, llm_row[0], llm_row[1] or '', llm_row[2] or 'gpt-4o-mini')
    updates = result.get('updates') or {}
    summary = result.get('summary') or '已记录'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO company_supplements (company_id, user_id, role, content, merged_fields) VALUES (?,?,?,?,?)",
        (company_id, current_user_id, 'user', text, None)
    )
    cursor.execute(
        "INSERT INTO company_supplements (company_id, user_id, role, content, merged_fields) VALUES (?,?,?,?,?)",
        (company_id, current_user_id, 'assistant', summary, json.dumps(updates, ensure_ascii=False))
    )
    set_clauses = ["last_updated=?"]
    params = [datetime.now().isoformat()]
    for key in ['legal_representative', 'registered_capital', 'business_status', 'registered_address', 'business_scope', 'social_evaluation', 'supplement_notes']:
        if key in updates and updates[key] is not None and str(updates[key]).strip():
            set_clauses.append(f"{key}=?")
            params.append(str(updates[key]).strip())
    if len(params) > 1:
        params.append(company_id)
        cursor.execute("UPDATE companies SET " + ", ".join(set_clauses) + " WHERE id=?", tuple(params))
    conn.commit()
    conn.close()
    return jsonify({'message': '已归集', 'summary': summary, 'updates': updates})


@app.route('/api/companies/<int:company_id>/ask', methods=['POST', 'OPTIONS'])
@token_required
def ask_company(current_user_id, company_id):
    """问这家企业：基于企业上下文（基本信息、新闻、尽调补充）用 LLM 回答用户问题。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    question = (data.get('question') or data.get('text') or '').strip()
    if not question:
        return jsonify({'message': '请输入问题'}), 400
    if len(question) > 500:
        return jsonify({'message': '问题不超过 500 字'}), 400
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM user_companies WHERE user_id=? AND company_id=?", (current_user_id, company_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '无权限操作该企业'}), 403
    cursor.execute("""SELECT name, industry, risk_level, legal_representative, registered_capital, business_status,
        registered_address, business_scope, social_evaluation, supplement_notes
        FROM companies WHERE id=?""", (company_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'message': 'Company not found'}), 404
    name = row[0]
    context_parts = [f"企业：{name}", f"行业：{row[1] or '-'}", f"风险等级：{row[2] or '-'}", f"法人：{row[3] or '-'}", f"注册资本：{row[4] or '-'}", f"经营状态：{row[5] or '-'}", f"社会评价：{row[8] or '-'}", f"尽调备注：{row[9] or '-'}"]
    cursor.execute("SELECT title, content, source, risk_level, category FROM company_news WHERE company_id=? ORDER BY created_at DESC LIMIT 30", (company_id,))
    for r in cursor.fetchall():
        context_parts.append(f"新闻：{r[0] or ''} | {r[2] or ''} | 风险:{r[3] or '-'} | {(r[1] or '')[:150]}")
    cursor.execute("SELECT role, content FROM company_supplements WHERE company_id=? ORDER BY created_at DESC LIMIT 20", (company_id,))
    for r in cursor.fetchall():
        context_parts.append(f"补充({r[0]}): {(r[1] or '')[:200]}")
    cursor.execute("SELECT api_key, base_url, model FROM llm_config WHERE user_id=?", (current_user_id,))
    llm_row = cursor.fetchone()
    conn.close()
    if not llm_row or not (llm_row[0] and str(llm_row[0]).strip()):
        return jsonify({'message': '请先在系统设置中配置 LLM API Key'}), 400
    from backend.services.llm_service import ask_company_question
    result = ask_company_question(name, '\n'.join(context_parts), question, llm_row[0], llm_row[1] or '', llm_row[2] or 'gpt-4o-mini')
    return jsonify(result)


@app.route('/api/settings/search-interval', methods=['GET', 'OPTIONS'])
@token_required
def get_search_interval(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT search_interval_minutes, report_template FROM user_settings WHERE user_id=?", (current_user_id,))
    row = cursor.fetchone()
    conn.close()
    return jsonify({'search_interval_minutes': row[0] if row else 30, 'report_template': row[1] if row else None})

@app.route('/api/settings/search-interval', methods=['PUT', 'OPTIONS'])
@token_required
def put_search_interval(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    interval = int(data.get('search_interval_minutes', 30))
    template = data.get('report_template', '')
    interval = max(10, min(1440, interval))
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM user_settings WHERE user_id=?", (current_user_id,))
    if cursor.fetchone():
        cursor.execute("UPDATE user_settings SET search_interval_minutes=?, report_template=?, updated_at=? WHERE user_id=?",
            (interval, template, datetime.now().isoformat(), current_user_id))
    else:
        cursor.execute("INSERT INTO user_settings (user_id, search_interval_minutes, report_template, updated_at) VALUES (?,?,?,?)",
            (current_user_id, interval, template, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'message': '设置已保存'})

def _analyze_and_classify_link(link_id, user_id, url, title, company_id):
    """后台分析链接并归类到企业、写入资讯"""
    db_path = _DB_PATH
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT api_key, base_url, model FROM llm_config WHERE user_id=?", (user_id,))
        llm_row = cur.fetchone()
        api_key, base_url, model = (llm_row[0], llm_row[1], llm_row[2]) if llm_row else (None, None, None)
        cur.execute("SELECT id, name FROM companies")
        companies_list = [(r[0], r[1]) for r in cur.fetchall()]
        conn.close()
        from backend.services.llm_service import analyze_document_or_link_for_company
        content = f"链接：{url}\n标题：{title}"
        result = analyze_document_or_link_for_company(title, content, companies_list, api_key, base_url, model)
        cid = result.get('company_id')
        if cid is None and company_id:
            cid = int(company_id)
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("UPDATE user_links SET company_id=?, analysis_result=?, status='analyzed' WHERE id=?",
            (cid, json.dumps(result, ensure_ascii=False), link_id))
        if result.get('is_news') and cid:
            rd = result.get('risk_dimensions') or {}
            cname_tag = next((c[1] for c in companies_list if c[0]==cid), '') or ''
            rl = result.get('risk_level', '')
            cur.execute("""INSERT INTO company_news (company_id, company_tag, title, content, source, source_url, sentiment_score, risk_level, category, keywords, publish_date, risk_dimensions)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cid, (cname_tag or '').strip(), title or url[:100], result.get('summary',''), '用户链接', url,
                 float(result.get('sentiment_score',0)), rl, result.get('category',''),
                 json.dumps(result.get('keywords',[]), ensure_ascii=False), datetime.now().strftime('%Y-%m-%d'),
                 json.dumps(rd, ensure_ascii=False)))
            if (rl or '').strip() == '高':
                _create_risk_alert(cur, cid, 'news', 'high', (title or url[:100] or '')[:200], '用户链接')
                _notify_alert_created(conn, cid, 'news', 'high', (title or url[:100] or '')[:200])
            cur.execute("INSERT INTO company_media_reviews (company_id, platform, title, content, sentiment) VALUES (?,?,?,?,?)",
                (cid, '用户链接', title or url[:100], result.get('summary',''), 'neutral' if float(result.get('sentiment_score',0))>=0 else 'negative'))
            if api_key:
                from backend.services.llm_service import analyze_sentiment_and_keywords
                cname = cname_tag or next((c[1] for c in companies_list if c[0]==cid), '')
                txt = (result.get('summary','') or '') + ' ' + ' '.join(result.get('keywords',[]))
                res = analyze_sentiment_and_keywords(cname, txt, api_key, base_url, model)
                cur.execute("UPDATE companies SET social_evaluation=?, risk_level=? WHERE id=?", (res.get('summary',''), res.get('risk_level',''), cid))
                if (res.get('risk_level') or '').strip() == '高':
                    _create_risk_alert(cur, cid, 'company', 'high', '企业风险等级升至高风险', 'link_analysis')
                    _notify_alert_created(conn, cid, 'company', 'high', '企业风险等级升至高风险')
                for kw in res.get('keywords', [])[:10]:
                    cur.execute("INSERT OR IGNORE INTO company_keywords (company_id, keyword, weight) VALUES (?,?,1)", (cid, kw))
        conn.commit()
        conn.close()
        if result.get('is_news') and cid:
            _sync_risk_pipeline_after_news(cid)
    except Exception as e:
        print('[Link] analyze error:', e)
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("UPDATE user_links SET status='failed', analysis_result=? WHERE id=?", (json.dumps({'error': str(e)}, ensure_ascii=False), link_id))
            conn.commit()
            conn.close()
        except: pass

@app.route('/api/links', methods=['POST', 'OPTIONS'])
@token_required
def add_link(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    url = data.get('url')
    company_id = data.get('company_id')
    if company_id is not None:
        company_id = int(company_id) if company_id else None
    title = data.get('title', '')
    if not url:
        return jsonify({'message': 'URL is required'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO user_links (user_id, company_id, url, title, status) VALUES (?,?,?,?,'pending')",
        (current_user_id, company_id, url, title))
    link_id = cursor.lastrowid
    conn.commit()
    conn.close()
    t = threading.Thread(target=_analyze_and_classify_link, args=(link_id, current_user_id, url, title, company_id))
    t.daemon = True
    t.start()
    return jsonify({'id': link_id, 'message': '链接已添加，正在自动分析归类'}), 201

@app.route('/api/links', methods=['GET', 'OPTIONS'])
@token_required
def get_links(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    company_id = request.args.get('company_id')
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    if company_id:
        cursor.execute("SELECT id, company_id, url, title, status, analysis_result, created_at FROM user_links WHERE user_id=? AND company_id=? ORDER BY created_at DESC",
            (current_user_id, company_id))
    else:
        cursor.execute("SELECT id, company_id, url, title, status, analysis_result, created_at FROM user_links WHERE user_id=? ORDER BY created_at DESC",
            (current_user_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'company_id': r[1], 'url': r[2], 'title': r[3], 'status': r[4], 'analysis_result': r[5], 'created_at': r[6]} for r in rows])

def _create_risk_alert(cursor, company_id, alert_type, severity, description, source=None):
    """在 risk_alerts 表中插入一条警报（企业风险升高或高风险新闻等）。"""
    try:
        cursor.execute("""INSERT INTO risk_alerts (company_id, alert_type, severity, description, source) VALUES (?,?,?,?,?)""",
            (company_id, alert_type, severity or 'high', description or '', source or ''))
    except sqlite3.OperationalError:
        pass


def _notify_alert_created(conn, company_id, alert_type, severity, description, company_name=None):
    """风险警报产生后：给收藏了该企业的用户创建站内通知，并根据预警规则发邮件。"""
    if not conn:
        return
    try:
        cur = conn.cursor()
        if company_name is None:
            cur.execute("SELECT name FROM companies WHERE id=?", (company_id,))
            r = cur.fetchone()
            company_name = r[0] if r else ''
        cur.execute("SELECT user_id FROM user_company_favorites WHERE company_id=?", (company_id,))
        user_ids = [row[0] for row in cur.fetchall()]
        title = f"风险警报：{company_name or ('企业#' + str(company_id))}"
        body = (description or '')[:500]
        rule_types_ok = ('new_alert', 'risk_upgrade')
        for uid in user_ids:
            cur.execute("SELECT id FROM alert_rules WHERE user_id=? AND enabled=1 AND rule_type IN (?,?)", (uid, 'new_alert', 'risk_upgrade'))
            if not cur.fetchone():
                continue
            cur.execute("""INSERT INTO notifications (user_id, type, title, body, related_type, related_id)
                VALUES (?,?,?,?,?,?)""", (uid, 'risk_alert', title, body, 'company', str(company_id)))
            send_email_to_user(uid, title, body + '\n\n请登录系统查看详情。')
        conn.commit()
    except Exception as e:
        logger.warning('_notify_alert_created: %s', e)


# Risk alerts routes
@app.route('/api/alerts', methods=['GET', 'OPTIONS'])
@token_required
def get_alerts(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    severity_filter = request.args.get('severity', '').strip().lower()
    conn = _db_conn()
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify([])
    placeholders = ','.join('?' * len(cids))
    sql = """
        SELECT ra.id, ra.company_id, ra.alert_type, ra.severity, ra.description, ra.source, ra.timestamp,
               c.name as company_name
        FROM risk_alerts ra
        JOIN companies c ON ra.company_id = c.id
        WHERE ra.company_id IN (""" + placeholders + """)
    """
    params = list(cids)
    try:
        sql = sql.replace("ra.timestamp,", "ra.timestamp, ra.processed_at,")
        sql += " ORDER BY ra.timestamp DESC LIMIT 50"
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()
        alerts = []
        for row in rows:
            alerts.append({
                'id': row[0],
                'company_id': row[1],
                'alert_type': row[2],
                'severity': row[3],
                'description': row[4],
                'source': row[5],
                'timestamp': row[6],
                'processed_at': row[7] if len(row) > 7 else None,
                'company_name': row[8] if len(row) > 8 else row[7]
            })
    except sqlite3.OperationalError:
        sql = """
            SELECT ra.id, ra.company_id, ra.alert_type, ra.severity, ra.description, ra.source, ra.timestamp,
                   c.name as company_name
            FROM risk_alerts ra
            JOIN companies c ON ra.company_id = c.id
            WHERE ra.company_id IN (""" + placeholders + """)
            ORDER BY ra.timestamp DESC
            LIMIT 50
        """
        cursor.execute(sql, tuple(cids))
        alerts = [{'id': row[0], 'company_id': row[1], 'alert_type': row[2], 'severity': row[3], 'description': row[4],
                   'source': row[5], 'timestamp': row[6], 'company_name': row[7], 'processed_at': None} for row in cursor.fetchall()]
    conn.close()
    if severity_filter and severity_filter in ('high', '中', '高', 'medium', '低', 'low'):
        sev_map = {'high': ('高', 'high'), 'medium': ('中', 'medium'), 'low': ('低', 'low'), '中': ('中', 'medium'), '高': ('高', 'high'), '低': ('低', 'low')}
        want_tuple = sev_map.get(severity_filter, (severity_filter,))
        alerts = [a for a in alerts if (a.get('severity') or '') in want_tuple or (a.get('severity') or '').lower() == severity_filter]
    return jsonify(alerts)


@app.route('/api/alerts/<int:alert_id>', methods=['PATCH', 'OPTIONS'])
@token_required
def mark_alert_processed(current_user_id, alert_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify({'message': '无权限'}), 403
    placeholders = ','.join('?' * len(cids))
    cursor.execute("SELECT id FROM risk_alerts WHERE id=? AND company_id IN (" + placeholders + ")", (alert_id,) + tuple(cids))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '警报不存在或无权限'}), 404
    try:
        cursor.execute("UPDATE risk_alerts SET processed_at=? WHERE id=?", (datetime.utcnow().isoformat(), alert_id))
    except sqlite3.OperationalError:
        conn.close()
        return jsonify({'message': '当前版本不支持标记已处理'}), 501
    conn.commit()
    conn.close()
    return jsonify({'message': '已标记为已处理'})

def _analyze_and_classify_document(doc_id, user_id, filepath, filename, company_id):
    """后台分析文档并归类到企业、写入资讯"""
    db_path = _DB_PATH
    try:
        from backend.services.document_service import extract_text_from_file
        content = extract_text_from_file(filepath)
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT api_key, base_url, model FROM llm_config WHERE user_id=?", (user_id,))
        llm_row = cur.fetchone()
        api_key, base_url, model = (llm_row[0], llm_row[1], llm_row[2]) if llm_row else (None, None, None)
        cur.execute("SELECT id, name FROM companies")
        companies_list = [(r[0], r[1]) for r in cur.fetchall()]
        conn.close()
        from backend.services.llm_service import analyze_document_or_link_for_company
        result = analyze_document_or_link_for_company(filename, content or filename, companies_list, api_key, base_url, model)
        cid = result.get('company_id')
        if cid is None and company_id:
            cid = int(company_id)
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("UPDATE documents SET company_id=?, analysis_result=?, status='analyzed' WHERE id=?",
            (cid, json.dumps(result, ensure_ascii=False), doc_id))
        if result.get('is_news') and cid:
            rd = result.get('risk_dimensions') or {}
            cname_tag = next((c[1] for c in companies_list if c[0]==cid), '') or ''
            rl = result.get('risk_level', '')
            cur.execute("""INSERT INTO company_news (company_id, company_tag, title, content, source, source_url, sentiment_score, risk_level, category, keywords, publish_date, risk_dimensions)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cid, (cname_tag or '').strip(), filename or '文档', result.get('summary','')[:1000], '用户上传', '',
                 float(result.get('sentiment_score',0)), rl, result.get('category',''),
                 json.dumps(result.get('keywords',[]), ensure_ascii=False), datetime.now().strftime('%Y-%m-%d'),
                 json.dumps(rd, ensure_ascii=False)))
            if (rl or '').strip() == '高':
                _create_risk_alert(cur, cid, 'news', 'high', (filename or '文档')[:200], '用户上传')
                _notify_alert_created(conn, cid, 'news', 'high', (filename or '文档')[:200])
            cur.execute("INSERT INTO company_media_reviews (company_id, platform, title, content, sentiment) VALUES (?,?,?,?,?)",
                (cid, '用户文档', filename, result.get('summary','')[:500], 'neutral' if float(result.get('sentiment_score',0))>=0 else 'negative'))
            if api_key:
                from backend.services.llm_service import analyze_sentiment_and_keywords
                cname = cname_tag or next((c[1] for c in companies_list if c[0]==cid), '')
                txt = (result.get('summary','') or '') + ' ' + ' '.join(result.get('keywords',[]))
                res = analyze_sentiment_and_keywords(cname, txt, api_key, base_url, model)
                cur.execute("UPDATE companies SET social_evaluation=?, risk_level=? WHERE id=?", (res.get('summary',''), res.get('risk_level',''), cid))
                if (res.get('risk_level') or '').strip() == '高':
                    _create_risk_alert(cur, cid, 'company', 'high', '企业风险等级升至高风险', 'document_analysis')
                    _notify_alert_created(conn, cid, 'company', 'high', '企业风险等级升至高风险')
                for kw in res.get('keywords', [])[:10]:
                    cur.execute("INSERT OR IGNORE INTO company_keywords (company_id, keyword, weight) VALUES (?,?,1)", (cid, kw))
        conn.commit()
        conn.close()
        if result.get('is_news') and cid:
            _sync_risk_pipeline_after_news(cid)
    except Exception as e:
        print('[Document] analyze error:', e)
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("UPDATE documents SET status='failed', analysis_result=? WHERE id=?", (json.dumps({'error': str(e)}, ensure_ascii=False), doc_id))
            conn.commit()
            conn.close()
        except: pass

# Document upload route
@app.route('/api/documents', methods=['POST', 'OPTIONS'])
@token_required
def upload_document(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
        
    if 'file' not in request.files:
        return jsonify({'message': 'No file provided'}), 400
    
    file = request.files['file']
    company_id = request.form.get('company_id')
    if company_id is not None and company_id != '':
        company_id = int(company_id)
    else:
        company_id = None
    
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    ext = (os.path.splitext(file.filename)[1] or '').lower()
    if ext not in ('.txt', '.pdf'):
        return jsonify({'message': '仅支持 .txt 或 .pdf 文件'}), 400
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > 10 * 1024 * 1024:
        return jsonify({'message': '文件大小不超过 10MB'}), 400

    from werkzeug.utils import secure_filename
    import uuid

    filename = secure_filename(file.filename)
    unique_filename = f"{uuid.uuid4()}_{filename}"
    upload_dir = os.path.join(_DATA_DIR, 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    filepath = os.path.join(upload_dir, unique_filename)
    
    file.save(filepath)
    
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO documents (user_id, company_id, filename, filepath, status)
        VALUES (?, ?, ?, ?, ?)
    """, (current_user_id, company_id, filename, filepath, 'pending'))
    doc_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    t = threading.Thread(target=_analyze_and_classify_document, args=(doc_id, current_user_id, filepath, filename, company_id))
    t.daemon = True
    t.start()
    
    return jsonify({
        'id': doc_id,
        'filename': filename,
        'filepath': filepath,
        'upload_time': datetime.now().isoformat(),
        'status': 'pending',
        'message': '文档已上传，正在自动分析归类'
    }), 201

# Risk indicators endpoint
@app.route('/api/risk-indicators', methods=['GET', 'OPTIONS'])
@token_required
def get_risk_indicators(current_user_id):
    """风险指标（按分类聚合），支持 ?days=7|30|90"""
    if request.method == 'OPTIONS':
        return jsonify({})
    days = min(90, max(1, int(request.args.get('days', 30))))
    conn = _db_conn()
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify([])
    placeholders = ','.join('?' * len(cids))
    cursor.execute("""SELECT category, risk_level, sentiment_score, COUNT(*) FROM company_news 
        WHERE company_id IN (""" + placeholders + """) AND created_at > datetime('now','-""" + str(days) + """ day') GROUP BY category, risk_level""", tuple(cids))
    rows = cursor.fetchall()
    conn.close()
    cat_levels = {}
    for cat, rl, sc, cnt in rows:
        key = cat or '其他'
        if key not in cat_levels:
            cat_levels[key] = {'low': 0, 'mid': 0, 'high': 0, 'count': 0, 'score_sum': 0}
        cat_levels[key]['count'] += cnt
        cat_levels[key]['score_sum'] += (float(sc or 0) * cnt)
        if rl == '低': cat_levels[key]['low'] += cnt
        elif rl == '中': cat_levels[key]['mid'] += cnt
        else: cat_levels[key]['high'] += cnt
    name_map = {'法律': '法律纠纷', '财务': '财务风险', '经营': '经营状况', '舆情': '舆情风险', '其他': '其他风险'}
    indicators = []
    for i, (cat, d) in enumerate(cat_levels.items()):
        total = d['count']
        level = '低' if d['high'] < total * 0.2 else ('高' if d['high'] > total * 0.5 else '中')
        avg = d['score_sum'] / total if total else 0
        change = f"{'+' if avg > 0 else ''}{avg*100:.0f}%"
        color = 'green' if level == '低' else ('red' if level == '高' else 'yellow')
        indicators.append({'id': i+1, 'name': name_map.get(cat, cat), 'level': level, 'change': change, 'color': color})
    # 无近期资讯时不返回占位数据，前端显示「暂无数据」；有数据时返回真实聚合指标
    return jsonify(indicators)


# ---------- 预警规则（可配置） ----------
@app.route('/api/alert-rules', methods=['GET', 'OPTIONS'])
@token_required
def list_alert_rules(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, rule_type, config, enabled, created_at FROM alert_rules WHERE user_id=? ORDER BY id", (current_user_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'name': r[1], 'rule_type': r[2], 'config': r[3], 'enabled': bool(r[4]), 'created_at': r[5]} for r in rows])


@app.route('/api/alert-rules', methods=['POST', 'OPTIONS'])
@token_required
def create_alert_rule(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    name = (data.get('name') or '').strip() or '未命名规则'
    rule_type = (data.get('rule_type') or 'risk_upgrade').strip()
    config = json.dumps(data.get('config') or {}, ensure_ascii=False)
    enabled = 1 if data.get('enabled', True) else 0
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO alert_rules (user_id, name, rule_type, config, enabled) VALUES (?,?,?,?,?)",
                   (current_user_id, name, rule_type, config, enabled))
    rid = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': rid, 'message': '规则已添加'}), 201


@app.route('/api/alert-rules/<int:rule_id>', methods=['PUT', 'PATCH', 'OPTIONS'])
@token_required
def update_alert_rule(current_user_id, rule_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM alert_rules WHERE id=? AND user_id=?", (rule_id, current_user_id))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '规则不存在或无权限'}), 404
    updates = []
    params = []
    if 'name' in data:
        updates.append('name=?')
        params.append((data.get('name') or '').strip() or '未命名规则')
    if 'rule_type' in data:
        updates.append('rule_type=?')
        params.append((data.get('rule_type') or 'risk_upgrade').strip())
    if 'config' in data:
        updates.append('config=?')
        params.append(json.dumps(data.get('config') or {}, ensure_ascii=False))
    if 'enabled' in data:
        updates.append('enabled=?')
        params.append(1 if data.get('enabled') else 0)
    if updates:
        params.append(rule_id)
        cursor.execute("UPDATE alert_rules SET " + ', '.join(updates) + " WHERE id=?", params)
        conn.commit()
    conn.close()
    return jsonify({'message': '已更新'})


@app.route('/api/alert-rules/<int:rule_id>', methods=['DELETE', 'OPTIONS'])
@token_required
def delete_alert_rule(current_user_id, rule_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM alert_rules WHERE id=? AND user_id=?", (rule_id, current_user_id))
    conn.commit()
    conn.close()
    return jsonify({'message': '已删除'})


# ---------- 站内通知 ----------
@app.route('/api/notifications', methods=['GET', 'OPTIONS'])
@token_required
def get_notifications(current_user_id):
    """站内通知列表，?limit=50&offset=0&unread_only=1"""
    if request.method == 'OPTIONS':
        return jsonify({})
    limit = min(100, max(1, int(request.args.get('limit', 50))))
    offset = max(0, int(request.args.get('offset', 0)))
    unread_only = request.args.get('unread_only', '').strip().lower() in ('1', 'true', 'yes')
    conn = _db_conn()
    cursor = conn.cursor()
    if unread_only:
        cursor.execute("""SELECT id, type, title, body, related_type, related_id, read_at, created_at
            FROM notifications WHERE user_id=? AND read_at IS NULL ORDER BY id DESC LIMIT ? OFFSET ?""",
                       (current_user_id, limit, offset))
    else:
        cursor.execute("""SELECT id, type, title, body, related_type, related_id, read_at, created_at
            FROM notifications WHERE user_id=? ORDER BY id DESC LIMIT ? OFFSET ?""",
                       (current_user_id, limit, offset))
    rows = cursor.fetchall()
    cursor.execute("SELECT COUNT(*) FROM notifications WHERE user_id=? AND read_at IS NULL", (current_user_id,))
    unread_count = cursor.fetchone()[0]
    conn.close()
    items = [{'id': r[0], 'type': r[1], 'title': r[2], 'body': r[3], 'related_type': r[4], 'related_id': r[5], 'read_at': r[6], 'created_at': r[7]} for r in rows]
    return jsonify({'items': items, 'unread_count': unread_count})


@app.route('/api/notifications/<int:nid>/read', methods=['PATCH', 'POST', 'OPTIONS'])
@token_required
def mark_notification_read(current_user_id, nid):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("UPDATE notifications SET read_at=? WHERE id=? AND user_id=?", (datetime.utcnow().isoformat(), nid, current_user_id))
    conn.commit()
    conn.close()
    return jsonify({'message': '已标为已读'})


@app.route('/api/notifications/read-all', methods=['PATCH', 'POST', 'OPTIONS'])
@token_required
def mark_all_notifications_read(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("UPDATE notifications SET read_at=? WHERE user_id=? AND read_at IS NULL", (datetime.utcnow().isoformat(), current_user_id))
    conn.commit()
    conn.close()
    return jsonify({'message': '已全部标为已读'})


# ---------- 操作审计日志 ----------
@app.route('/api/audit-log', methods=['GET', 'OPTIONS'])
@token_required
def get_audit_log(current_user_id):
    """查询审计日志：普通用户仅本人，管理员可查全部。?limit=50&offset=0&user_id=（admin 可选）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    limit = min(200, max(1, int(request.args.get('limit', 50))))
    offset = max(0, int(request.args.get('offset', 0)))
    conn = _db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id=?", (current_user_id,))
    row = cursor.fetchone()
    is_admin = row and row[0] == 'admin'
    filter_user = request.args.get('user_id')
    if is_admin and filter_user:
        cursor.execute("""SELECT a.id, a.user_id, u.username, a.action, a.resource_type, a.resource_id, a.detail, a.ip, a.created_at
            FROM audit_log a LEFT JOIN users u ON u.id = a.user_id WHERE a.user_id=? ORDER BY a.id DESC LIMIT ? OFFSET ?""",
                       (int(filter_user), limit, offset))
    elif is_admin:
        cursor.execute("""SELECT a.id, a.user_id, u.username, a.action, a.resource_type, a.resource_id, a.detail, a.ip, a.created_at
            FROM audit_log a LEFT JOIN users u ON u.id = a.user_id ORDER BY a.id DESC LIMIT ? OFFSET ?""", (limit, offset))
    else:
        cursor.execute("""SELECT a.id, a.user_id, u.username, a.action, a.resource_type, a.resource_id, a.detail, a.ip, a.created_at
            FROM audit_log a LEFT JOIN users u ON u.id = a.user_id WHERE a.user_id=? ORDER BY a.id DESC LIMIT ? OFFSET ?""",
                       (current_user_id, limit, offset))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'user_id': r[1], 'username': r[2], 'action': r[3], 'resource_type': r[4], 'resource_id': r[5], 'detail': r[6], 'ip': r[7], 'created_at': r[8]} for r in rows])


# ---------- 企业批量导入 ----------
@app.route('/api/companies/batch-import', methods=['POST', 'OPTIONS'])
@token_required
def batch_import_companies(current_user_id):
    """批量添加企业。body: { "companies": [ { "name": "企业名", "industry": "行业" }, ... ] }，最多 100 条"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    items = data.get('companies') or []
    if not isinstance(items, list) or len(items) > 100:
        return jsonify({'message': '请提供 companies 数组，最多 100 条'}), 400
    conn = _db_conn()
    cursor = conn.cursor()
    created = []
    for item in items:
        name = (item.get('name') or '').strip()
        if not name:
            continue
        industry = (item.get('industry') or '').strip()
        cursor.execute("INSERT INTO companies (name, industry, crawl_status) VALUES (?, ?, 'pending')", (name, industry))
        cid = cursor.lastrowid
        cursor.execute("INSERT OR IGNORE INTO user_companies (user_id, company_id) VALUES (?, ?)", (current_user_id, cid))
        created.append({'id': cid, 'name': name, 'industry': industry})
    conn.commit()
    conn.close()
    _audit_log(current_user_id, 'batch_import', 'company', None, f'导入{len(created)}家')
    # 为每家新企业自动启动与「添加企业」相同的流程：工商信息 + 相关新闻 + 媒体舆情爬取
    for c in created:
        t = threading.Thread(target=_run_company_crawl_and_news_pipeline, args=(c['id'], c['name'], current_user_id))
        t.daemon = True
        t.start()
    return jsonify({'message': f'已导入 {len(created)} 家企业，后台正在自动爬取工商信息、相关新闻与媒体舆情', 'created': created}), 201


# ---------- 备份与恢复 ----------
@app.route('/api/backup', methods=['POST', 'OPTIONS'])
@token_required
@admin_required
def create_backup(current_user_id):
    """立即创建数据库备份（仅管理员）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    backup_dir = os.path.join(os.path.dirname(_DB_PATH), 'backup')
    os.makedirs(backup_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(backup_dir, f'risk_platform_{ts}.db')
    try:
        import shutil
        shutil.copy2(_DB_PATH, dest)
        _audit_log(current_user_id, 'backup', 'system', None, dest)
        return jsonify({'message': '备份已创建', 'path': dest, 'filename': os.path.basename(dest)})
    except Exception as e:
        return jsonify({'message': str(e)}), 500


@app.route('/api/backup/list', methods=['GET', 'OPTIONS'])
@token_required
@admin_required
def list_backups(current_user_id):
    """列出备份文件（仅管理员）"""
    if request.method == 'OPTIONS':
        return jsonify({})
    backup_dir = os.path.join(os.path.dirname(_DB_PATH), 'backup')
    if not os.path.isdir(backup_dir):
        return jsonify([])
    files = []
    for f in os.listdir(backup_dir):
        if f.endswith('.db'):
            path = os.path.join(backup_dir, f)
            try:
                mtime = os.path.getmtime(path)
                files.append({'filename': f, 'path': path, 'modified': datetime.fromtimestamp(mtime).isoformat()})
            except OSError:
                pass
    files.sort(key=lambda x: x['modified'], reverse=True)
    return jsonify(files[:50])


@app.route('/api/backup/restore', methods=['POST', 'OPTIONS'])
@token_required
@admin_required
def restore_backup(current_user_id):
    """从备份恢复（仅管理员）。body: { "filename": "risk_platform_xxx.db" }"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    filename = (data.get('filename') or '').strip()
    if not filename or '..' in filename or '/' in filename:
        return jsonify({'message': '无效的 filename'}), 400
    backup_dir = os.path.join(os.path.dirname(_DB_PATH), 'backup')
    path = os.path.join(backup_dir, filename)
    if not os.path.isfile(path):
        return jsonify({'message': '备份文件不存在'}), 404
    try:
        import shutil
        shutil.copy2(path, _DB_PATH)
        _audit_log(current_user_id, 'restore_backup', 'system', None, filename)
        return jsonify({'message': '已恢复，请重启后端使数据生效'})
    except Exception as e:
        return jsonify({'message': str(e)}), 500


# ---------- 任务执行记录（细化） ----------
@app.route('/api/task-status/detailed', methods=['GET', 'OPTIONS'])
@token_required
def get_task_runs_detailed(current_user_id):
    """最近任务执行记录：新闻搜索、爬取等。仅返回当前用户有权限的企业相关任务。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    limit = min(100, max(1, int(request.args.get('limit', 30))))
    conn = _db_conn()
    cursor = conn.cursor()
    cids = _user_company_ids(cursor, current_user_id)
    if not cids:
        conn.close()
        return jsonify([])
    placeholders = ','.join('?' * len(cids))
    cursor.execute("""
        SELECT t.id, t.task_type, t.company_id, t.status, t.message, t.started_at, t.finished_at, c.name
        FROM task_runs t
        LEFT JOIN companies c ON c.id = t.company_id
        WHERE t.company_id IN (""" + placeholders + """)
        ORDER BY t.id DESC
        LIMIT ?
    """, tuple(cids) + (limit,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'task_type': r[1], 'company_id': r[2], 'status': r[3], 'message': r[4], 'started_at': r[5], 'finished_at': r[6], 'company_name': r[7]} for r in rows])


# ---------- API 文档（OpenAPI） ----------
@app.route('/api/openapi.yaml', methods=['GET'])
def serve_openapi():
    """提供 OpenAPI 3.0 规范文件"""
    import os as _os
    p = _os.path.join(_os.path.dirname(__file__), 'openapi.yaml')
    if not _os.path.isfile(p):
        return jsonify({'message': 'openapi.yaml not found'}), 404
    from flask import send_file
    return send_file(p, mimetype='application/x-yaml', as_attachment=False)


@app.route('/api/docs', methods=['GET'])
def api_docs_ui():
    """Swagger UI 页面（通过 CDN 加载 openapi.yaml）"""
    from flask import Response
    base = request.host_url.rstrip('/')
    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"/><title>API 文档</title>
<link rel="stylesheet" href="https://unpkg.com/swagger-ui-dist@5/swagger-ui.css"/></head>
<body><div id="swagger-ui"></div>
<script src="https://unpkg.com/swagger-ui-dist@5/swagger-ui-bundle.js"></script>
<script>
  SwaggerUIBundle({{ url: "{base}/api/openapi.yaml", dom_id: "#swagger-ui" }});
</script></body></html>'''
    return Response(html, mimetype='text/html')


# ---------- 健康检查（含依赖状态） ----------
@app.route('/api/health/detailed', methods=['GET'])
def health_detailed():
    """健康检查：db、可选 LLM 配置、MediaCrawler 路径"""
    out = {'status': 'ok', 'db': 'ok', 'llm_configured': False, 'mediacrawler_path': bool(os.environ.get('MEDIACRAWLER_PATH'))}
    try:
        c = _db_conn()
        c.cursor().execute('SELECT 1').fetchone()
        c.close()
    except Exception as e:
        out['status'] = 'degraded'
        out['db'] = str(e)[:200]
    try:
        conn = _db_conn()
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM llm_config WHERE api_key IS NOT NULL AND api_key != '' LIMIT 1")
        out['llm_configured'] = cur.fetchone() is not None
        conn.close()
    except Exception:
        pass
    return jsonify(out)


# Admin routes (require admin role)
@app.route('/api/admin/users', methods=['GET', 'OPTIONS'])
@token_required
@admin_required
def get_users(current_user_id):
    if request.method == 'OPTIONS':
        return jsonify({})
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, email, role, created_at FROM users ORDER BY created_at DESC")
    users = []
    for row in cursor.fetchall():
        users.append({
            'id': row[0],
            'username': row[1],
            'email': row[2],
            'role': row[3],
            'created_at': row[4]
        })
    
    conn.close()
    return jsonify(users)


@app.route('/api/admin/users', methods=['POST', 'OPTIONS'])
@token_required
@admin_required
def admin_create_user(current_user_id):
    """管理员创建用户（可指定角色）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    email = (data.get('email') or '').strip()
    password = (data.get('password') or '').strip()
    role = (data.get('role') or 'user').strip().lower()
    if role not in ('user', 'admin'):
        role = 'user'
    if not username:
        return jsonify({'message': '请填写用户名'}), 400
    if not password or len(password) < 6:
        return jsonify({'message': '密码至少 6 位'}), 400
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users (username, email, password_hash, role) VALUES (?, ?, ?, ?)",
                       (username, email or None, generate_password_hash(password), role))
        conn.commit()
        uid = cursor.lastrowid
        conn.close()
        return jsonify({'message': '创建成功', 'id': uid}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'message': '用户名或邮箱已存在'}), 400


@app.route('/api/admin/users/<int:user_id>', methods=['PATCH', 'PUT', 'OPTIONS'])
@token_required
@admin_required
def admin_update_user(current_user_id, user_id):
    """管理员修改用户（角色、邮箱等）。不能修改自己的角色为普通用户（防止锁死）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    if user_id == current_user_id:
        return jsonify({'message': '不能修改自己的角色'}), 400
    data = request.get_json() or {}
    conn = sqlite3.connect(_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '用户不存在'}), 404
    updates = []
    params = []
    if 'role' in data and data['role'] in ('user', 'admin'):
        updates.append("role=?")
        params.append(data['role'])
    if 'email' in data:
        updates.append("email=?")
        params.append((data['email'] or '').strip() or None)
    if 'username' in data and (data['username'] or '').strip():
        updates.append("username=?")
        params.append((data['username'] or '').strip())
    if not updates:
        conn.close()
        return jsonify({'message': '无有效修改'}), 400
    params.append(user_id)
    cursor.execute("UPDATE users SET " + ", ".join(updates) + " WHERE id=?", tuple(params))
    conn.commit()
    conn.close()
    return jsonify({'message': '已更新'})


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE', 'OPTIONS'])
@token_required
@admin_required
def admin_delete_user(current_user_id, user_id):
    """管理员删除用户及其关联数据。不能删除自己。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    if user_id == current_user_id:
        return jsonify({'message': '不能删除自己'}), 400
    db_path = _DB_PATH
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'message': '用户不存在'}), 404
    for table, col in [
        ('user_company_favorites', 'user_id'),
        ('llm_config', 'user_id'),
        ('user_settings', 'user_id'),
        ('documents', 'user_id'),
        ('user_links', 'user_id'),
        ('news_items', 'user_id'),
        ('risk_insights', 'user_id'),
    ]:
        try:
            cursor.execute("DELETE FROM " + table + " WHERE " + col + "=?", (user_id,))
        except sqlite3.OperationalError:
            pass
    try:
        cursor.execute("DELETE FROM company_supplements WHERE user_id=?", (user_id,))
    except sqlite3.OperationalError:
        pass
    cursor.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '已删除'})


@app.route('/api/admin/system-settings', methods=['GET', 'PUT', 'PATCH', 'OPTIONS'])
@token_required
@admin_required
def admin_system_settings(current_user_id):
    """管理员获取或更新系统设置（如 SMTP 发件人）。GET 返回当前配置（密码脱敏），PUT/PATCH 更新。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    db_path = _DB_PATH
    if request.method in ('PUT', 'PATCH'):
        data = request.get_json() or {}
        allowed = ('smtp_host', 'smtp_port', 'smtp_user', 'smtp_from', 'smtp_password')
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        for key in allowed:
            if key not in data:
                continue
            val = data[key]
            if val is None or (isinstance(val, str) and val.strip() == ''):
                cursor.execute("DELETE FROM system_settings WHERE key=?", (key,))
            else:
                cursor.execute("INSERT OR REPLACE INTO system_settings (key, value) VALUES (?, ?)", (key, str(val).strip()))
        conn.commit()
        conn.close()
        return jsonify({'message': '已保存'})
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM system_settings WHERE key LIKE 'smtp_%'")
    rows = cursor.fetchall()
    conn.close()
    out = {'smtp_host': '', 'smtp_port': '587', 'smtp_user': '', 'smtp_from': '', 'smtp_password': ''}
    for k, v in rows:
        if k in out and v is not None:
            if k == 'smtp_password' and v:
                out[k] = '********'
            else:
                out[k] = str(v).strip()
    if not out.get('smtp_host') and os.environ.get('SMTP_HOST'):
        out['smtp_host'] = os.environ.get('SMTP_HOST', '')
        out['smtp_port'] = str(os.environ.get('SMTP_PORT', '587'))
        out['smtp_user'] = os.environ.get('SMTP_USER', '')
        out['smtp_from'] = (os.environ.get('SMTP_FROM') or os.environ.get('SMTP_USER') or '')
        out['smtp_password'] = '********' if os.environ.get('SMTP_PASSWORD') else ''
    return jsonify(out)


@app.route('/api/admin/risk-labels', methods=['GET', 'POST', 'OPTIONS'])
@token_required
@admin_required
def admin_risk_labels(current_user_id):
    """
    管理企业风险标签：
    - GET: 查询标签（可按 enterprise_id / label_type 过滤）
    - POST: 写入或更新标签（INSERT OR REPLACE）
    """
    if request.method == 'OPTIONS':
        return jsonify({})
    db_path = _DB_PATH
    if request.method == 'GET':
        enterprise_id = request.args.get('enterprise_id')
        label_type = request.args.get('label_type')
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        sql = "SELECT id, enterprise_id, as_of_date, label_type, label_value, note, created_at FROM enterprise_risk_label WHERE 1=1"
        params = []
        if enterprise_id:
            sql += " AND enterprise_id=?"
            params.append(int(enterprise_id))
        if label_type:
            sql += " AND label_type=?"
            params.append(label_type.strip())
        sql += " ORDER BY as_of_date DESC, created_at DESC"
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()
        conn.close()
        out = []
        for r in rows:
            out.append({
                'id': r[0],
                'enterprise_id': r[1],
                'as_of_date': r[2],
                'label_type': r[3],
                'label_value': r[4],
                'note': r[5],
                'created_at': r[6],
            })
        return jsonify(out)

    # POST: 创建/更新标签
    data = request.get_json() or {}
    enterprise_id = data.get('enterprise_id')
    as_of_date = (data.get('as_of_date') or '').strip()
    label_type = (data.get('label_type') or 'explosion').strip()
    label_value = data.get('label_value')
    note = (data.get('note') or '').strip() or None

    try:
        enterprise_id = int(enterprise_id)
    except (TypeError, ValueError):
        return jsonify({'message': 'enterprise_id 必须是数字'}), 400
    if not as_of_date or len(as_of_date) < 10:
        return jsonify({'message': 'as_of_date 必须是 YYYY-MM-DD'}), 400
    try:
        val = int(label_value)
    except (TypeError, ValueError):
        return jsonify({'message': 'label_value 必须是 0 或 1'}), 400
    if val not in (0, 1):
        return jsonify({'message': 'label_value 必须是 0 或 1'}), 400

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT OR REPLACE INTO enterprise_risk_label
            (id, enterprise_id, as_of_date, label_type, label_value, note, created_at)
        VALUES (
            (SELECT id FROM enterprise_risk_label WHERE enterprise_id=? AND as_of_date=? AND label_type=?),
            ?, ?, ?, ?, ?, COALESCE(
                (SELECT created_at FROM enterprise_risk_label WHERE enterprise_id=? AND as_of_date=? AND label_type=?),
                CURRENT_TIMESTAMP
            )
        )
        """,
        (
            enterprise_id, as_of_date, label_type,
            enterprise_id, as_of_date, label_type, val, note,
            enterprise_id, as_of_date, label_type,
        ),
    )
    conn.commit()
    conn.close()
    _audit_log(
        current_user_id,
        'admin_set_risk_label',
        resource_type='enterprise',
        resource_id=enterprise_id,
        detail=f'{label_type}={val} on {as_of_date}',
    )
    return jsonify({'message': '已保存标签'})


def _update_company_risk_level_from_news(conn, company_id):
    """
    根据该企业 company_news 的风险等级汇总更新 companies.risk_level，便于首页/详情页显示且无社会评价时不为「未知」。
    若汇总结果为「高」且原等级非高，则写入一条风险警报。
    """
    if conn is None:
        conn = _db_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT risk_level FROM company_news WHERE company_id=? AND risk_level IN ('高','中','低')",
        (company_id,),
    )
    levels = [row[0].strip() for row in cur.fetchall() if row and row[0]]
    if not levels:
        return
    order = {'高': 3, '中': 2, '低': 1}
    derived = max(levels, key=lambda x: order.get(x, 0))
    cur.execute("SELECT risk_level FROM companies WHERE id=?", (company_id,))
    row = cur.fetchone()
    old_level = (row[0] or '').strip() if row else ''
    cur.execute("UPDATE companies SET risk_level=? WHERE id=?", (derived, company_id))
    if derived == '高' and old_level != '高':
        _create_risk_alert(cur, company_id, 'company', 'high', '企业相关新闻存在高风险，已更新风险等级', 'news_aggregate')
    conn.commit()


def _sync_risk_pipeline_after_news(company_id=None):
    """company_news 有新数据写入后，自动更新该企业（或全库）的事件特征与时间序列，并依新闻汇总更新企业风险等级。"""
    try:
        from backend.services import feature_extraction_service as fes
        from backend.services import risk_timeseries_service as rts
        fes.DB_PATH = _DB_PATH
        rts.DB_PATH = _DB_PATH
        fes.rebuild_news_features(db_path=_DB_PATH, enterprise_id=company_id)
        if company_id is not None:
            rts.rebuild_timeseries_for_enterprise(company_id, db_path=_DB_PATH)
            conn = _db_conn()
            try:
                _update_company_risk_level_from_news(conn, company_id)
            finally:
                conn.close()
        else:
            conn = _db_conn()
            try:
                cur = conn.cursor()
                cur.execute("SELECT id FROM companies")
                cids = [row[0] for row in cur.fetchall()]
                for cid in cids:
                    rts.rebuild_timeseries_for_enterprise(cid, db_path=_DB_PATH)
                    _update_company_risk_level_from_news(conn, cid)
            finally:
                conn.close()
    except Exception as e:
        logger.exception('_sync_risk_pipeline_after_news error: %s', e)


@app.route('/api/admin/rebuild_features', methods=['POST', 'OPTIONS'])
@token_required
@admin_required
def admin_rebuild_features(current_user_id):
    """管理员触发：从 company_news 重建企业事件特征（enterprise_event_feature）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    enterprise_id = data.get('enterprise_id')
    try:
        from backend.services import feature_extraction_service as fes
        fes.DB_PATH = _DB_PATH
        count = fes.rebuild_news_features(enterprise_id=enterprise_id)
        conn = _db_conn()
        try:
            if enterprise_id is not None:
                _update_company_risk_level_from_news(conn, int(enterprise_id))
            else:
                cur = conn.cursor()
                cur.execute("SELECT id FROM companies")
                for row in cur.fetchall():
                    _update_company_risk_level_from_news(conn, row[0])
        finally:
            conn.close()
        _audit_log(
            current_user_id,
            'admin_rebuild_features',
            resource_type='enterprise' if enterprise_id else 'all',
            resource_id=enterprise_id,
            detail=f'news_features={count}',
        )
        return jsonify({'message': '重建完成', 'news_event_count': count})
    except Exception as e:
        logger.exception('admin_rebuild_features error')
        return jsonify({'message': '重建失败', 'error': str(e)}), 500


@app.route('/api/admin/db/tables', methods=['GET', 'OPTIONS'])
@token_required
@admin_required
def admin_db_tables(current_user_id):
    """管理员：列出 SQLite 中所有表名（仅 table 类型，排除 sqlite_sequence 等）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    try:
        conn = _db_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        )
        tables = [row[0] for row in cur.fetchall()]
        conn.close()
        return jsonify({'tables': tables})
    except Exception as e:
        logger.exception('admin_db_tables error')
        return jsonify({'message': str(e), 'tables': []}), 500


def _safe_table_name(name):
    """只允许字母、数字、下划线，防止 SQL 注入。"""
    if not name or not isinstance(name, str):
        return None
    s = name.strip()
    return s if s and s.replace('_', '').isalnum() else None


@app.route('/api/admin/db/table/<table_name>', methods=['GET', 'OPTIONS'])
@token_required
@admin_required
def admin_db_table_data(current_user_id, table_name):
    """管理员：查看指定表的结构与数据（分页，只读）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    safe_name = _safe_table_name(table_name)
    if not safe_name:
        return jsonify({'message': '无效表名'}), 400
    limit = min(500, max(1, request.args.get('limit', type=int) or 100))
    offset = max(0, request.args.get('offset', type=int) or 0)
    try:
        conn = _db_conn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (safe_name,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'message': '表不存在'}), 404
        cur.execute("PRAGMA table_info(%s)" % safe_name)
        columns = [{'name': row[1], 'type': row[2], 'pk': bool(row[5])} for row in cur.fetchall()]
        cur.execute("SELECT COUNT(*) FROM %s" % safe_name)
        total = cur.fetchone()[0]
        cur.execute("SELECT * FROM %s LIMIT ? OFFSET ?" % safe_name, (limit, offset))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({
            'table': safe_name,
            'columns': columns,
            'total': total,
            'limit': limit,
            'offset': offset,
            'rows': rows,
        })
    except Exception as e:
        logger.exception('admin_db_table_data error')
        return jsonify({'message': str(e)}), 500


def _get_table_columns_and_pk(cursor, safe_name):
    """返回 (columns_list, pk_column_name)。pk_column_name 可能为 None。"""
    cursor.execute("PRAGMA table_info(%s)" % safe_name)
    rows = cursor.fetchall()
    cols = [row[1] for row in rows]
    pk = None
    for row in rows:
        if row[5]:  # pk
            pk = row[1]
            break
    return cols, pk


@app.route('/api/admin/db/table/<table_name>/row', methods=['POST', 'PUT', 'DELETE', 'OPTIONS'])
@token_required
@admin_required
def admin_db_table_row(current_user_id, table_name):
    """管理员：增改删单行。POST=插入，PUT=按主键更新，DELETE=按主键删除。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    safe_name = _safe_table_name(table_name)
    if not safe_name:
        return jsonify({'message': '无效表名'}), 400
    try:
        conn = _db_conn()
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (safe_name,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'message': '表不存在'}), 404
        columns, pk = _get_table_columns_and_pk(cur, safe_name)
        data = request.get_json() or {}

        if request.method == 'POST':
            # 插入：只接受表中存在的列，主键若为 AUTOINCREMENT 可省略
            row = data.get('row') if isinstance(data.get('row'), dict) else data
            cols = [c for c in row.keys() if c in columns]
            if not cols:
                conn.close()
                return jsonify({'message': '请提供至少一列有效字段'}), 400
            placeholders = ','.join('?' * len(cols))
            col_list = ','.join(cols)
            cur.execute(
                "INSERT INTO %s (%s) VALUES (%s)" % (safe_name, col_list, placeholders),
                [row[c] for c in cols]
            )
            conn.commit()
            rid = cur.lastrowid
            conn.close()
            return jsonify({'message': '已插入', 'id': rid})

        if request.method == 'PUT':
            if not pk:
                conn.close()
                return jsonify({'message': '该表无主键，无法按行更新'}), 400
            pk_val = data.get('id') or data.get(pk)
            if pk_val is None:
                conn.close()
                return jsonify({'message': '请提供主键值 (id 或 %s)' % pk}), 400
            row = data.get('row') if isinstance(data.get('row'), dict) else data
            cols = [c for c in row.keys() if c in columns and c != pk]
            if not cols:
                conn.close()
                return jsonify({'message': '请提供要更新的列'}), 400
            set_clause = ','.join("%s=?" % c for c in cols)
            cur.execute(
                "UPDATE %s SET %s WHERE %s=?" % (safe_name, set_clause, pk),
                [row[c] for c in cols] + [pk_val]
            )
            conn.commit()
            conn.close()
            return jsonify({'message': '已更新', 'rows': cur.rowcount})

        if request.method == 'DELETE':
            if not pk:
                conn.close()
                return jsonify({'message': '该表无主键，无法按行删除'}), 400
            pk_val = data.get('id') or data.get(pk)
            if pk_val is None:
                conn.close()
                return jsonify({'message': '请提供主键值 (id 或 %s)' % pk}), 400
            cur.execute("DELETE FROM %s WHERE %s=?" % (safe_name, pk), (pk_val,))
            conn.commit()
            conn.close()
            return jsonify({'message': '已删除', 'rows': cur.rowcount})
    except Exception as e:
        logger.exception('admin_db_table_row error')
        return jsonify({'message': str(e)}), 500


@app.route('/api/admin/rebuild_risk_timeseries', methods=['POST', 'OPTIONS'])
@token_required
@admin_required
def admin_rebuild_risk_timeseries(current_user_id):
    """管理员触发：从 enterprise_event_feature 重建企业风险时间序列（enterprise_risk_timeseries）。"""
    if request.method == 'OPTIONS':
        return jsonify({})
    data = request.get_json() or {}
    enterprise_id = data.get('enterprise_id')
    try:
        from backend.services import risk_timeseries_service as rts
        rts.DB_PATH = _DB_PATH
        total = 0
        if enterprise_id:
            total = rts.rebuild_timeseries_for_enterprise(int(enterprise_id))
        else:
            conn = sqlite3.connect(_DB_PATH)
            cur = conn.cursor()
            cur.execute("SELECT id FROM companies")
            ids = [row[0] for row in cur.fetchall()]
            conn.close()
            for cid in ids:
                total += rts.rebuild_timeseries_for_enterprise(cid)
        _audit_log(
            current_user_id,
            'admin_rebuild_risk_timeseries',
            resource_type='enterprise' if enterprise_id else 'all',
            resource_id=enterprise_id,
            detail=f'days_written={total}',
        )
        return jsonify({'message': '重建完成', 'days_written': total})
    except Exception as e:
        logger.exception('admin_rebuild_risk_timeseries error')
        return jsonify({'message': '重建失败', 'error': str(e)}), 500

# Media crawler integration routes (simplified version)
try:
    from simple_media_integration import crawl_company_risks_simple, simple_media_crawler
    print("Simple media crawler integration loaded successfully")
    
    @app.route('/api/media-crawl/company/<company_name>', methods=['GET', 'OPTIONS'])
    @token_required
    def crawl_company_for_risks(current_user_id, company_name):
        if request.method == 'OPTIONS':
            return jsonify({})
            
        """为企业爬取社交媒体上的风险信息"""
        if not crawl_company_risks_simple:
            return jsonify({'error': 'Media crawler not available'}), 500
            
        try:
            # 在后台线程中运行异步爬虫
            def run_crawl():
                return asyncio.run(crawl_company_risks_simple(company_name))
            
            # 使用线程池执行异步函数
            future = ThreadPoolExecutor(max_workers=1).submit(run_crawl)
            results = future.result(timeout=300)  # 5分钟超时
            
            return jsonify(results)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/media-crawl/configure', methods=['POST', 'OPTIONS'])
    @token_required
    def configure_media_crawler(current_user_id):
        if request.method == 'OPTIONS':
            return jsonify({})
            
        """配置媒体爬虫参数"""
        return jsonify({
            'message': 'Configuration not needed for simple crawler',
            'config': {
                'status': 'simple_crawler_active'
            }
        })

except ImportError:
    pass  # 使用内置爬虫 services/crawler.py
    
    @app.route('/api/media-crawl/company/<company_name>', methods=['GET', 'OPTIONS'])
    @token_required
    def crawl_company_for_risks_stub(current_user_id, company_name):
        if request.method == 'OPTIONS':
            return jsonify({})
        return jsonify({'error': 'Media crawler not available'}), 500
    
    @app.route('/api/media-crawl/configure', methods=['POST', 'OPTIONS'])
    @token_required
    def configure_media_crawler_stub(current_user_id):
        if request.method == 'OPTIONS':
            return jsonify({})
        return jsonify({'error': 'Media crawler not available'}), 500


def _standalone_static_root():
    """独立运行/打包时前端静态目录。打包为 .app 时 exe 在 Contents/MacOS、资源在 Contents/Frameworks。"""
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(os.path.abspath(getattr(sys, 'executable', __file__)))
        # 若在 .app 内（.../xxx.app/Contents/MacOS），前端在 ../Frameworks/frontend
        frameworks = os.path.normpath(os.path.join(exe_dir, '..', 'Frameworks'))
        frontend_in_frameworks = os.path.join(frameworks, 'frontend')
        if os.path.isdir(frontend_in_frameworks):
            return os.path.abspath(frontend_in_frameworks)
        return os.path.abspath(os.path.join(getattr(sys, '_MEIPASS', exe_dir), 'frontend'))
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), 'risk_frontend', 'build')


_standalone_mode = os.environ.get('STANDALONE_APP', '').strip() in ('1', 'true', 'yes') or getattr(sys, 'frozen', False)
if _standalone_mode:
    # 让 Flask 内置的 /static/ 从打包后的前端 static 目录提供，否则会先于我们的路由返回 404
    _frontend_root = _standalone_static_root()
    _static_dir = os.path.join(os.path.abspath(_frontend_root), 'static')
    if os.path.isdir(_static_dir):
        app.static_folder = _static_dir
        app.static_url_path = '/static'
    _MIME_OVERRIDES = {'.js': 'application/javascript', '.css': 'text/css', '.json': 'application/json'}
    @app.route('/api/debug-static-root')
    def debug_static_root():
        """调试：返回当前静态根路径及是否存在（仅 standalone 模式）"""
        root = _standalone_static_root()
        return jsonify({
            'root': root,
            'exists': os.path.isdir(root),
            'index_exists': os.path.isfile(os.path.join(root, 'index.html')) if root else False,
        })
    # 显式处理 /static/ 避免被其他逻辑误判为 404
    @app.route('/static/<path:subpath>')
    def serve_static(subpath):
        from flask import Response
        root = os.path.abspath(_standalone_static_root())
        if not os.path.isdir(root):
            return jsonify({'error': 'root not found', 'root': root}), 404
        full = os.path.normpath(os.path.join(root, 'static', subpath))
        root_real = os.path.realpath(root)
        if not (full == root_real or full.startswith(root_real + os.sep)):
            return jsonify({'error': 'invalid path', 'root': root, 'full': full}), 404
        if not os.path.isfile(full):
            return jsonify({
                'error': 'file not found',
                'root': root,
                'full': full,
                'list_dir': os.listdir(root)[:15] if os.path.isdir(root) else [],
            }), 404
        _, ext = os.path.splitext(subpath)
        mimetype = _MIME_OVERRIDES.get(ext.lower(), 'application/octet-stream')
        try:
            with open(full, 'rb') as f:
                return Response(f.read(), mimetype=mimetype)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    @app.route('/', defaults={'path': ''})
    @app.route('/<path:path>')
    def serve_frontend(path):
        from flask import Response
        root = os.path.abspath(_standalone_static_root())
        if not os.path.isdir(root):
            return 'Frontend build not found. Run: npm run build:frontend', 404
        if path:
            path = path.lstrip('/').replace('..', '')
        if path:
            full = os.path.normpath(os.path.join(root, path))
            # 用 realpath 统一符号链接，避免 root/full 不一致导致 startswith 失败
            root_real = os.path.realpath(root)
            full_real = os.path.realpath(full) if os.path.exists(full) else full
            if not full_real.startswith(root_real):
                return 'Invalid path', 404
            if os.path.isfile(full):
                _, ext = os.path.splitext(path)
                mimetype = _MIME_OVERRIDES.get(ext.lower(), 'application/octet-stream')
                try:
                    with open(full, 'rb') as f:
                        return Response(f.read(), mimetype=mimetype)
                except Exception as e:
                    logger.exception('serve_frontend read error')
                    return jsonify({'error': str(e)}), 500
            # 文件不存在时返回调试信息（仅 /static/ 且 frozen）
            if path.startswith('static/') and getattr(sys, 'frozen', False):
                try:
                    ls = os.listdir(root)
                    static_ls = os.listdir(os.path.join(root, 'static')) if os.path.isdir(os.path.join(root, 'static')) else []
                except Exception:
                    ls = static_ls = []
                return jsonify({
                    'debug': 'file not found',
                    'root': root,
                    'path': path,
                    'full': full,
                    'root_exists': os.path.isdir(root),
                    'list_root': ls[:20],
                    'list_static': static_ls[:20],
                }), 404
        index_path = os.path.join(root, 'index.html')
        if os.path.isfile(index_path):
            try:
                with open(index_path, 'rb') as f:
                    return Response(f.read(), mimetype='text/html; charset=utf-8')
            except Exception as e:
                logger.exception('serve_frontend index error')
                return jsonify({'error': str(e)}), 500
        return 'Frontend index.html not found', 404


def run_standalone_server():
    """独立运行时的服务入口（init_db、定时任务、Flask）。供桌面窗口启动器在子线程中调用。"""
    if not os.environ.get('SKIP_DROP_TABLES') and os.environ.get('FLASK_ENV') != 'production':
        logger.warning(
            '未设置 SKIP_DROP_TABLES=1 或 FLASK_ENV=production，'
            '若为生产环境可能导致数据库被清空，请务必在 .env 或环境中配置。'
        )
    init_db()
    try:
        from backend.services.scheduler_service import start_scheduler
        conn = sqlite3.connect(_DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT search_interval_minutes FROM user_settings WHERE search_interval_minutes > 0 LIMIT 1")
        row = cur.fetchone()
        conn.close()
        interval = int(row[0]) if row and row[0] else 30
        interval = max(10, min(1440, interval))
        start_scheduler(interval)
    except Exception as e:
        logger.warning('Scheduler start error: %s', e)
    debug = not _is_production
    port = int(os.environ.get('PORT', 8005))
    host = os.environ.get('HOST') or ('127.0.0.1' if _standalone_mode else '0.0.0.0')
    if _standalone_mode and not os.environ.get('DESKTOP_WINDOW', '').strip() in ('1', 'true', 'yes'):
        def _open_browser():
            time.sleep(1.5)
            import webbrowser
            webbrowser.open(f'http://127.0.0.1:{port}/')
        threading.Thread(target=_open_browser, daemon=True).start()
    app.run(debug=debug, port=port, host=host, use_reloader=False)


if __name__ == '__main__':
    _secret = app.config.get('SECRET_KEY') or ''
    if _secret == 'your-secret-key-change-in-production' or (isinstance(_secret, str) and len(_secret) < 20):
        logger.warning(
            'SECRET_KEY 仍为默认或过短，生产环境请设置强随机密钥。'
            '可运行: python backend/scripts/generate_secret_key.py'
        )
    run_standalone_server()